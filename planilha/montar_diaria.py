# -*- coding: utf-8 -*-
"""
Programação Diária — Makro Transportes
======================================
Caderno de campo para o tablet. O controle de verdade é o PROTHEUS: aqui só se
anota o que tem de ser feito, o número da OS que foi aberta lá, e se saiu ou
não. Uma aba de trabalho e uma de listas.

A regra que manda no desenho: **nenhuma fórmula nas linhas.** Só quatro contas,
no topo. Isso deixa a planilha leve no tablet, faz apagar e inserir linha ser
seguro, e tira qualquer chance de #REF!.

    python3 montar_diaria.py dados.json

`dados.json` vem do ler_planilha.py rodado sobre a planilha antiga.
"""
import sys, json, io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.comments import Comment

SAIDA="/home/user/gestaomkt/planilha/Programacao_Diaria_Makro.xlsx"

NAVY="FF13164E"; NAVY2="FF2A3170"; BRANCO="FFFFFFFF"
TINTA="FF14181F"; T2="FF4A5464"; BORDA="FFC8D0DE"; CINZA="FFEEF1F7"
AMARELO="FFFFF3CF"
OK_V="FFDCF3E4"; OK_T="FF116B3E"
AL_V="FFFFE9CC"; AL_T="FF9A4C00"
RU_V="FFFDE4E7"; RU_T="FFB00020"

F=lambda **k: Font(name="Arial", **k)
fina=Side(style="thin", color=BORDA)
box=Border(left=fina,right=fina,top=fina,bottom=fina)
def fill(c): return PatternFill("solid", fgColor=c)
DFMT='[$-416]dd/mm/yyyy;@'

CAB=6                      # linha do cabeçalho da tabela
PRIM=CAB+1                 # primeira linha de dados
ULT=706                    # 700 linhas: sobra muita e não pesa, pois não há fórmula
MCAB=8; MPRIM=MCAB+1; MULT=MPRIM+299        # aba de movimentações, 300 linhas

# ═══════════════════════════════════ entrada
args=[a for a in sys.argv[1:] if not a.startswith("--")]
FONTE=json.load(io.open(args[0],encoding="utf-8")) if args else {"linhas":[],"listas":{}}
LST=FONTE.get("listas",{})

def d2(s):
    if not s: return None
    if isinstance(s,date): return s
    a,m,d=str(s)[:10].split("-"); return date(int(a),int(m),int(d))

# As movimentações são digitadas AQUI, não vêm da analítica. Sem isto, toda
# reconstrução apagaria o que ele lançou. Lê de volta do próprio arquivo.
MOVS=[]
try:
    from openpyxl import load_workbook as _lw
    import os as _os
    if _os.path.exists(SAIDA):
        _w=_lw(SAIDA, data_only=True)
        if "Movimentações" in _w.sheetnames:
            _m=_w["Movimentações"]
            for _r in range(MPRIM,MULT+1):
                if not _m[f"A{_r}"].value: continue
                MOVS.append([_m[f"{c}{_r}"].value for c in ("A","B","C","D","E","H","I")])
except Exception as _e:
    sys.stderr.write(f"aviso: não consegui reler as movimentações ({_e})\n")

# ── migração: o formato antigo vira o novo ──
# Cancelada não vem: no caderno de campo ela só atrapalha. Fica na planilha
# antiga, que continua no repositório.
LINHAS=[]
for x in FONTE["linhas"]:
    if x.get("marcar")=="Cancelada": continue
    if not (x.get("atividade") or x.get("frota")): continue
    obs=" · ".join(t for t in (x.get("motivo",""),x.get("obs","")) if t)
    LINHAS.append(dict(
        # O PRAZO é o dia da atividade; o Início é o dia em que o serviço
        # inteiro começa. É o prazo que manda no dia a dia.
        data=d2(x.get("prazo") or x.get("data")),
        frota=x.get("frota",""),
        atividade=x.get("atividade",""),
        feito=("Sim" if (x.get("concluida") or x.get("marcar")=="Concluída") else ""),
        os=x.get("os",""),
        quem=" · ".join(x.get("equipe") or []),
        servico=x.get("servico",""),
        obs=obs))
# O PENDENTE primeiro, e dentro dele por data: o topo da planilha é sempre o
# que falta fazer. O que já saiu desce para o fim, em cinza. Sem data vai para
# depois do que tem dia marcado.
LINHAS.sort(key=lambda l:(l["feito"]=="Sim",
                          l["data"] or date(2099,1,1), l["frota"], l["servico"]))

FROTAS=sorted(set([f for f in LST.get("Frotas",[]) if f] +
                  [l["frota"] for l in LINHAS if l["frota"]]))
QUEM=[q for q in LST.get("Executantes",[]) if q]

# ═══════════════════════════════════ DIA A DIA
wb=Workbook()
ws=wb.active; ws.title="Dia a dia"

ws["A1"]="PROGRAMAÇÃO DIÁRIA  ·  MAKRO TRANSPORTES"
ws["A1"].font=F(bold=True,size=15,color=BRANCO); ws["A1"].fill=fill(NAVY)
ws["A1"].alignment=Alignment(vertical="center",indent=1)
ws.merge_cells("A1:H1"); ws.row_dimensions[1].height=32

ws["A2"]=("O controle é o PROTHEUS. Aqui você anota a atividade, o número da OS "
          "que abriu lá, e marca se saiu.")
ws["A2"].font=F(size=10,italic=True,color=T2)
ws["A2"].alignment=Alignment(vertical="center",indent=1)
ws.merge_cells("A2:H2"); ws.row_dimensions[2].height=18

# ── a data que se olha ──
ws["A3"]="DIA"
ws["A3"].font=F(bold=True,size=12,color=NAVY)
ws["A3"].alignment=Alignment(horizontal="right",vertical="center")
ws["B3"]="=TODAY()"
d=ws["B3"]; d.font=F(bold=True,size=13,color=NAVY); d.fill=fill(AMARELO)
d.border=Border(*[Side(style="medium",color=NAVY)]*4)
d.alignment=Alignment(horizontal="center",vertical="center"); d.number_format=DFMT
d.comment=Comment("Vem no dia de hoje.\n\nTroque a data para ver outro dia — os "
 "números do lado acompanham. Para voltar, escreva =HOJE()","PCM")
ws["C3"]='=IF($B$3="","",TEXT($B$3,"[$-416]dddd"))'
ws["C3"].font=F(bold=True,size=12,color=T2)
ws["C3"].alignment=Alignment(horizontal="center",vertical="center")
ws.row_dimensions[3].height=30

# ── as quatro contas, e só elas ──
DATA=f"$A${PRIM}:$A${ULT}"; ATIV=f"$C${PRIM}:$C${ULT}"
FEITO=f"$D${PRIM}:$D${ULT}"; OSC=f"$E${PRIM}:$E${ULT}"
PROG=f'COUNTIFS({DATA},$B$3,{ATIV},"<>")'
FEIT=f'COUNTIFS({DATA},$B$3,{ATIV},"<>",{FEITO},"Sim")'
IND=[("D","E","PROGRAMADAS",f"={PROG}","0",NAVY,
      "atividades com esta data"),
     ("F","F","FEITAS",f"={FEIT}","0",OK_T,
      "dessas, quantas você marcou Sim"),
     ("G","G","ADERÊNCIA DO DIA",'=IF($D$4=0,"",$F$4/$D$4)',"0%",NAVY,
      "feitas ÷ programadas do dia"),
     ("H","H","SEM OS",f'={PROG}-COUNTIFS({DATA},$B$3,{ATIV},"<>",{OSC},"<>")',"0",AL_T,
      "atividades de hoje que ainda não têm OS aberta no Protheus")]
for c1,c2,rot,fml,fmt,cor,dica in IND:
    if c1!=c2: ws.merge_cells(f"{c1}3:{c2}3"); ws.merge_cells(f"{c1}4:{c2}4")
    a=ws[f"{c1}3"]; a.value=rot
    a.font=F(bold=True,size=8.5,color=T2); a.fill=fill(CINZA)
    a.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    a.comment=Comment(dica,"PCM")
    b=ws[f"{c1}4"]; b.value=fml
    b.font=F(bold=True,size=18,color=cor); b.number_format=fmt
    b.alignment=Alignment(horizontal="center",vertical="center"); b.fill=fill(CINZA)
    for cc in {c1,c2}:
        ws[f"{cc}3"].border=box; ws[f"{cc}4"].border=box
ws.row_dimensions[4].height=30
# atrasadas e carteira numa linha só, sem virar mais um quadrado
ws["A5"]=(f'="Atrasadas, de dias anteriores:  "&'
          f'(COUNTIFS({DATA},"<"&$B$3,{ATIV},"<>")-'
          f'COUNTIFS({DATA},"<"&$B$3,{ATIV},"<>",{FEITO},"Sim"))&'
          f'"        |        Sem data, esperando entrar na programação:  "&'
          # contar vazio com COUNTIFS(...,"") não é confiável entre engines;
          # o total menos as que têm data é inequívoco.
          f'(COUNTIFS({ATIV},"<>")-COUNTIFS({ATIV},"<>",{DATA},"<>"))')
ws["A5"].font=F(size=10,bold=True,color=T2)
ws["A5"].alignment=Alignment(vertical="center",indent=1)
ws.merge_cells("A5:H5"); ws.row_dimensions[5].height=22

# ── a tabela ──
# Ordem pensada para o tablet: o que se lê e o que se toca fica nas cinco
# primeiras colunas, que cabem numa tela. Quem faz, serviço e observação
# ficam à direita, para quando precisar.
COLS=[("A","Data",11),("B","Frota",13),("C","Atividade",44),
      ("D","Feito",9),("E","OS Protheus",13),
      ("F","Quem faz",17),("G","Serviço",20),("H","Obs.",22)]
DICAS={"C":"O que tem de ser feito. Escreva à vontade.",
       "D":"Toque e escolha SIM quando sair. Vazio é pendente.",
       "E":"O número da OS aberta no Protheus.\n\nVazio quer dizer que ainda "
           "não foi aberta — é assim que se sabe o que falta lançar lá.",
       "A":"O dia em que a atividade está programada.\n\nDeixe vazio para ela "
           "ficar esperando, sem dia — ela some do quadro de cima e fica na "
           "conta de “sem data”."}
for letra,tit,larg in COLS:
    c=ws[f"{letra}{CAB}"]; c.value=tit
    c.font=F(bold=True,size=11,color=BRANCO); c.fill=fill(NAVY2)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=box
    ws.column_dimensions[letra].width=larg
    if letra in DICAS: c.comment=Comment(DICAS[letra],"PCM")
ws.row_dimensions[CAB].height=26

for i in range(ULT-PRIM+1):
    r=PRIM+i
    l=LINHAS[i] if i<len(LINHAS) else None
    if l:
        for letra,val in (("A",l["data"]),("B",l["frota"]),("C",l["atividade"]),
                          ("D",l["feito"]),("E",l["os"]),("F",l["quem"]),
                          ("G",l["servico"]),("H",l["obs"])):
            if val: ws[f"{letra}{r}"]=val
    for letra,tit,larg in COLS:
        c=ws[f"{letra}{r}"]
        # Fonte 11 e linha alta: no tablet o dedo precisa de alvo.
        c.font=F(size=11,color=TINTA); c.border=box
        if letra=="A": c.number_format=DFMT
        if letra=="E": c.number_format="@"      # OS não pode perder zero à esquerda
        if letra in ("A","B","D","E"):
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            c.alignment=Alignment(vertical="center",indent=1)
    ws.row_dimensions[r].height=21

# ── caixas de seleção ──
def dv(col, formula, titulo, msg, brando=True):
    v=DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    v.errorTitle=titulo; v.error=msg; v.showErrorMessage=True
    if brando: v.errorStyle="warning"
    ws.add_data_validation(v); v.add(f"{col}{PRIM}:{col}{ULT}")
LIN_L=max(len(FROTAS),len(QUEM))+40
dv("B", f"=Listas!$A$2:$A${LIN_L}", "Frota",
   "Frota fora da lista. Pode continuar — depois acrescente na aba Listas.")
dv("F", f"=Listas!$B$2:$B${LIN_L}", "Quem faz",
   "Nome fora da lista. Pode continuar — depois acrescente na aba Listas.")
dv("D", '"Sim,Não"', "Feito", "Escolha Sim ou Não, ou deixe vazio.", brando=False)

# ── cores: só o que ajuda a decidir ──
FX=f"A{PRIM}:H{ULT}"
def regra(faixa, formula, **est):
    r=Rule(type="expression", dxf=DifferentialStyle(**est)); r.formula=[formula]
    ws.conditional_formatting.add(faixa, r)
# feita: a célula fica verde
regra(f"D{PRIM}:D{ULT}", f'$D{PRIM}="Sim"',
      fill=PatternFill(bgColor=OK_V), font=Font(color=OK_T,bold=True))
regra(f"D{PRIM}:D{ULT}", f'$D{PRIM}="Não"',
      fill=PatternFill(bgColor=RU_V), font=Font(color=RU_T,bold=True))
# o dia chegou e não saiu: a data acende
regra(f"A{PRIM}:A{ULT}",
      f'AND($C{PRIM}<>"",$A{PRIM}<>"",$A{PRIM}<=TODAY(),$D{PRIM}<>"Sim")',
      fill=PatternFill(bgColor=AL_V), font=Font(color=AL_T,bold=True))
# sem OS no Protheus, tendo dia marcado: é o que falta lançar lá
regra(f"E{PRIM}:E{ULT}", f'AND($C{PRIM}<>"",$A{PRIM}<>"",$E{PRIM}="")',
      fill=PatternFill(bgColor=AL_V))
# linha inteira apagada quando já saiu, para o olho ir no que falta
regra(FX, f'$D{PRIM}="Sim"', font=Font(color="FF98A1B2"))

ws.freeze_panes=f"A{PRIM}"
ws.auto_filter.ref=f"A{CAB}:H{ULT}"
ws.sheet_view.showGridLines=False

# ═══════════════════════════════════ LISTAS
ls=wb.create_sheet("Listas")
ls["A1"]="Frotas"; ls["B1"]="Quem faz"
for c in ("A","B"):
    x=ls[f"{c}1"]; x.font=F(bold=True,size=11,color=BRANCO)
    x.fill=fill(NAVY2); x.alignment=Alignment(horizontal="center"); x.border=box
    ls.column_dimensions[c].width=22
for i in range(LIN_L-1):
    r=2+i
    for c,vals in (("A",FROTAS),("B",QUEM)):
        cc=ls[f"{c}{r}"]
        if i<len(vals): cc.value=vals[i]
        cc.font=F(size=11); cc.border=box
        cc.alignment=Alignment(horizontal="center")
    ls.row_dimensions[r].height=19
ls["D1"]="COMO USAR"
ls["D1"].font=F(bold=True,size=12,color=NAVY)
AJUDA=["Acrescente nomes aqui e eles aparecem nas caixinhas da aba Dia a dia.",
       "",
       "NO DIA A DIA:",
       "1.  A caixa amarela lá em cima é o DIA. Vem em hoje; troque para ver outro.",
       "2.  Para achar o dia: toque na setinha da coluna Data e filtre.",
       "3.  Fez? Toque em FEITO e escolha Sim. A linha apaga e o número sobe.",
       "4.  Abriu a OS no Protheus? Escreva o número em OS PROTHEUS.",
       "     Enquanto estiver vazia a célula fica laranja — é o que falta lançar lá.",
       "",
       "PARA ACRESCENTAR ATIVIDADE:",
       "     Escreva numa linha vazia no fim. Data, frota, atividade, e pronto.",
       "     Sem data ela fica esperando, e entra na conta de “sem data”.",
       "",
       "Não há fórmula nenhuma nas linhas: pode apagar, inserir e arrastar à",
       "vontade que nada quebra."]
for i,t in enumerate(AJUDA):
    c=ls[f"D{2+i}"]; c.value=t
    c.font=F(size=10, bold=t.endswith(":"), color=NAVY if t.endswith(":") else TINTA)
    c.alignment=Alignment(vertical="center")
ls.column_dimensions["D"].width=76
ls.sheet_view.showGridLines=False

# ═══════════════════════════════════ MOVIMENTAÇÕES
# A operação leva a frota até o fornecedor, ou traz de volta. Quando ela
# demora, a oficina fica com a atividade vencida e o motivo vira "frota não
# chegou". Esta aba mede essa demora: quem prometeu, para quando, e quando
# chegou de verdade. É o número para cobrar, no lugar da conversa de memória.
mv=wb.create_sheet("Movimentações")
mv["A1"]="MOVIMENTAÇÃO DE FROTAS  ·  o que a operação prometeu e quando entregou"
mv["A1"].font=F(bold=True,size=15,color=BRANCO); mv["A1"].fill=fill(NAVY)
mv["A1"].alignment=Alignment(vertical="center",indent=1)
mv.merge_cells("A1:I1"); mv.row_dimensions[1].height=32
mv["A2"]=("Peça a movimentação, anote a data que a operação prometeu, e marque "
          "quando a frota chegou. O atraso e os números saem sozinhos.")
mv["A2"].font=F(size=10,italic=True,color=T2)
mv["A2"].alignment=Alignment(vertical="center",indent=1)
mv.merge_cells("A2:I2"); mv.row_dimensions[2].height=18
# única célula volátil da aba
mv["L1"]="=TODAY()"; mv["L1"].font=F(size=8,color=T2); mv["L1"].number_format=DFMT
mv["K1"]="hoje — não apague"; mv["K1"].font=F(bold=True,size=8,color=RU_T)
mv.column_dimensions["K"].width=16; mv.column_dimensions["L"].width=11

MFROTA=f"$A${MPRIM}:$A${MULT}"; MPROM=f"$D${MPRIM}:$D${MULT}"
MCHEG=f"$E${MPRIM}:$E${MULT}"; MATR=f"$F${MPRIM}:$F${MULT}"
ENTR=f'COUNTIFS({MFROTA},"<>",{MCHEG},"<>")'
ATRASADAS=f'COUNTIFS({MFROTA},"<>",{MATR},">0")'
MIND=[("A","B","MOVIMENTAÇÕES",f'=COUNTIFS({MFROTA},"<>")',"0",NAVY,
       "tudo que foi pedido à operação"),
      ("C","C","ENTREGUES",f"={ENTR}","0",OK_T,"já chegaram"),
      ("D","D","NO PRAZO",f"={ENTR}-{ATRASADAS}","0",OK_T,
       "chegaram até a data prometida"),
      ("E","F","PONTUALIDADE DA OPERAÇÃO",'=IF($C$4=0,"",$D$4/$C$4)',"0%",NAVY,
       "no prazo ÷ entregues. É este o número para levar à reunião"),
      ("G","H","ATRASO MÉDIO",
       f'=IF({ATRASADAS}=0,"",SUMIFS({MATR},{MFROTA},"<>")/{ATRASADAS})',"0.0",AL_T,
       "dias, contando só as que atrasaram"),
      ("I","I","DIAS PERDIDOS",f'=SUMIFS({MATR},{MFROTA},"<>")',"0",RU_T,
       "a soma de todos os atrasos: o tamanho do prejuízo em dias de oficina")]
for c1,c2,rot,fml,fmt,cor,dica in MIND:
    if c1!=c2: mv.merge_cells(f"{c1}3:{c2}3"); mv.merge_cells(f"{c1}4:{c2}4")
    a_=mv[f"{c1}3"]; a_.value=rot
    a_.font=F(bold=True,size=8.5,color=T2); a_.fill=fill(CINZA)
    a_.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    a_.comment=Comment(dica,"PCM")
    b_=mv[f"{c1}4"]; b_.value=fml
    b_.font=F(bold=True,size=18,color=cor); b_.number_format=fmt
    b_.alignment=Alignment(horizontal="center",vertical="center"); b_.fill=fill(CINZA)
    for cc in {c1,c2}: mv[f"{cc}3"].border=box; mv[f"{cc}4"].border=box
mv.row_dimensions[3].height=24; mv.row_dimensions[4].height=30
mv.conditional_formatting.add("E4", CellIsRule(operator="greaterThanOrEqual",
    formula=["0.85"], font=F(bold=True,size=18,color=OK_T)))
mv.conditional_formatting.add("E4", CellIsRule(operator="lessThan",
    formula=["0.6"], font=F(bold=True,size=18,color=RU_T)))
mv[f"A6"]=("Atraso = quando chegou menos quando foi prometida. Chegar antes conta "
           "como zero, não como crédito.")
mv["A6"].font=F(size=9,italic=True,color=T2)
mv["A6"].alignment=Alignment(vertical="center",indent=1)
mv.merge_cells("A6:I6"); mv.row_dimensions[6].height=18

MCOLS=[("A","Frota",11),("B","Destino / fornecedor",26),("C","Pedida em",12),
       ("D","Prometida para",13),("E","Chegou em",12),("F","Atraso",9),
       ("G","Situação",17),("H","Para quê",30),("I","Quem prometeu",18)]
MDICAS={"C":"O dia em que você pediu a movimentação à operação.",
        "D":"A data que a operação SE COMPROMETEU a entregar. É contra ela que "
            "o atraso é medido — sem essa data não há o que cobrar.",
        "E":"O dia em que a frota chegou de verdade. Enquanto estiver vazio, "
            "a movimentação está em aberto.",
        "F":"Sai sozinho: chegou menos prometida, em dias. Chegar antes conta zero.",
        "I":"Quem, na operação, deu a data. Sem nome, a cobrança vira conversa."}
for letra,tit,larg in MCOLS:
    c=mv[f"{letra}{MCAB}"]; c.value=tit
    c.font=F(bold=True,size=11,color=BRANCO); c.fill=fill(NAVY2)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=box
    mv.column_dimensions[letra].width=larg
    if letra in MDICAS: c.comment=Comment(MDICAS[letra],"PCM")
mv.row_dimensions[MCAB].height=26

for i in range(MULT-MPRIM+1):
    r=MPRIM+i
    if i<len(MOVS):
        for letra,val in zip(("A","B","C","D","E","H","I"),MOVS[i]):
            if val not in (None,""): mv[f"{letra}{r}"]=val
    # As duas fórmulas olham só a própria linha: apagar linha não quebra nada.
    mv[f"F{r}"]=f'=IF(OR($A{r}="",$D{r}="",$E{r}=""),"",MAX(0,$E{r}-$D{r}))'
    mv[f"G{r}"]=(f'=IF($A{r}="","",IF($E{r}<>"",IF(N($F{r})>0,"Entregue com atraso",'
                 f'"Entregue no prazo"),IF($D{r}="","Sem data prometida",'
                 f'IF($D{r}<$L$1,"ATRASADA","Aguardando"))))')
    for letra,tit,larg in MCOLS:
        c=mv[f"{letra}{r}"]
        c.font=F(size=11,color=TINTA); c.border=box
        if letra in ("C","D","E"): c.number_format=DFMT
        if letra in ("A","C","D","E","F","G"):
            c.alignment=Alignment(horizontal="center",vertical="center")
        else: c.alignment=Alignment(vertical="center",indent=1)
    mv.row_dimensions[r].height=21

dvm=DataValidation(type="list", formula1=f"=Listas!$A$2:$A${LIN_L}",
                   allow_blank=True, showDropDown=False)
dvm.errorTitle="Frota"; dvm.error="Frota fora da lista — pode continuar."
dvm.errorStyle="warning"; dvm.showErrorMessage=True
mv.add_data_validation(dvm); dvm.add(f"A{MPRIM}:A{MULT}")

MFX=f"A{MPRIM}:I{MULT}"
def mregra(faixa, formula, **est):
    r_=Rule(type="expression", dxf=DifferentialStyle(**est)); r_.formula=[formula]
    mv.conditional_formatting.add(faixa, r_)
for txt,bg,fg in (("Entregue no prazo",OK_V,OK_T),("Entregue com atraso",AL_V,AL_T),
                  ("ATRASADA",RU_V,RU_T),("Aguardando","FFEDF2FC","FF2C4272"),
                  ("Sem data prometida","FFF6E8CE","FF8A5A05")):
    mregra(f"G{MPRIM}:G{MULT}", f'$G{MPRIM}="{txt}"',
           fill=PatternFill(bgColor=bg), font=Font(color=fg,bold=True))
mregra(f"F{MPRIM}:F{MULT}", f'N($F{MPRIM})>0',
       fill=PatternFill(bgColor=RU_V), font=Font(color=RU_T,bold=True))
mregra(f"D{MPRIM}:D{MULT}", f'AND($A{MPRIM}<>"",$D{MPRIM}<>"",$E{MPRIM}="",$D{MPRIM}<$L$1)',
       fill=PatternFill(bgColor=RU_V), font=Font(color=RU_T,bold=True))
mregra(MFX, f'$E{MPRIM}<>""', font=Font(color="FF98A1B2"))

mv.freeze_panes=f"A{MPRIM}"
mv.auto_filter.ref=f"A{MCAB}:I{MULT}"
mv.sheet_view.showGridLines=False

# ═══════════════════════════════════ FECHO
for w,orient in ((ws,"landscape"),(ls,"portrait"),(mv,"landscape")):
    w.page_setup.orientation=orient
    w.page_setup.paperSize=w.PAPERSIZE_A4
    w.page_setup.fitToWidth=1; w.page_setup.fitToHeight=0
    w.sheet_properties.pageSetUpPr.fitToPage=True
    w.page_margins.left=w.page_margins.right=0.4
    w.page_margins.top=w.page_margins.bottom=0.5
ws.print_title_rows=f"{CAB}:{CAB}"
mv.print_title_rows=f"{MCAB}:{MCAB}"
mv.print_area=f"A1:I{MPRIM+max(len(MOVS),25)+3}"
ws.print_area=f"A1:H{PRIM+len(LINHAS)+4}"
wb.calculation.fullCalcOnLoad=True
for w in wb.worksheets: w.sheet_properties.tabColor=NAVY[2:]
wb.active=0
wb.save(SAIDA)
print("planilha montada:",SAIDA)
print(f"  atividades: {len(LINHAS)} | com data: {sum(1 for l in LINHAS if l['data'])}"
      f" | feitas: {sum(1 for l in LINHAS if l['feito']=='Sim')}"
      f" | com OS: {sum(1 for l in LINHAS if l['os'])}")
print(f"  frotas: {len(FROTAS)} | executantes: {len(QUEM)} | colunas: {len(COLS)}")
print(f"  movimentações preservadas: {len(MOVS)}")
