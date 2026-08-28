# -*- coding: utf-8 -*-
"""
Lê uma planilha já em uso e devolve as linhas em JSON, para o
montar_planilha.py reconstruir a estrutura sem perder nada do que foi
digitado.

    python3 ler_planilha.py Programacao.xlsx > dados_atuais.json

Lê pelo TÍTULO da coluna na linha 3, e não pela letra, então sobrevive a
qualquer troca de posição. Reconhece o layout por número de semana e, como
reserva, o antigo por data.
"""
import sys, json, io
from openpyxl import load_workbook

def texto(v):
    if v is None: return ""
    return str(v).strip()

def num(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None

def iso(v):
    return v.date().isoformat() if hasattr(v,"date") else (v.isoformat() if hasattr(v,"isoformat") else "")

def ler(caminho):
    wb=load_workbook(caminho, data_only=True)
    pg=wb["Programação"]
    # descobre onde está cada coluna pelo título, e não pela letra: assim a
    # leitura sobrevive a qualquer troca de posição
    col={}
    for c in range(1,60):
        t=texto(pg.cell(row=3,column=c).value)
        if t: col.setdefault(t,c)
    def v(r,titulo):
        c=col.get(titulo)
        return pg.cell(row=r,column=c).value if c else None

    linhas=[]
    for r in range(4, pg.max_row+1):
        ativ=texto(v(r,"Atividade"))
        # Linha reservada — frota já lançada, serviço ainda por escrever —
        # também tem de sobreviver ao ida e volta. Só se descarta a linha que
        # não tem nem atividade nem frota, que é linha realmente vazia.
        if not ativ and not texto(v(r,"Frota")): continue
        eq=[texto(v(r,f"Executante {i}")) for i in (1,2,3)]
        linhas.append(dict(
            os=texto(v(r,"OS")), frota=texto(v(r,"Frota")),
            servico=texto(v(r,"Serviço")), atividade=ativ,
            tipo=texto(v(r,"Tipo")) or "Corretiva",
            origem=texto(v(r,"Origem")),
            equipe=[x for x in eq if x],
            # Programar virou NÚMERO de semana + dia; "Data programada" e
            # "1ª data" são do layout antigo e ficam como reserva, senão a
            # leitura de uma planilha nova devolvia semana e dia vazios e o
            # ida e volta apagava a programação inteira.
            semana=num(v(r,"Semana")), dia=texto(v(r,"Dia")),
            semorig=num(v(r,"Semana orig.")),
            dias=num(v(r,"Dias prev.")),
            data=iso(v(r,"Início")) or iso(v(r,"Data programada")),
            concluida=iso(v(r,"Concluída em")),
            marcar=texto(v(r,"Marcar")),
            orig=iso(v(r,"1ª data")),
            motivo=texto(v(r,"Motivo")), obs=texto(v(r,"Obs.")),
            oficina=texto(v(r,"Oficina")),
        ))
    # listas de apoio, para não perder frotas/executantes cadastrados à mão
    listas={}
    if "Listas" in wb.sheetnames:
        ls=wb["Listas"]
        for c in range(1,10):
            t=texto(ls.cell(row=4,column=c).value)
            if not t: continue
            vals=[]
            for rr in range(5,200):
                x=ls.cell(row=rr,column=c).value
                if x is None: continue
                if hasattr(x,"date"): continue      # a coluna de semanas some
                vals.append(texto(x))
            if vals: listas[t]=vals
    return {"linhas":linhas,"listas":listas}

if __name__=="__main__":
    print(json.dumps(ler(sys.argv[1]), ensure_ascii=False, indent=1))
