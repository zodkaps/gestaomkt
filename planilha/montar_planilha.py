# -*- coding: utf-8 -*-
"""
Programação de Serviços — Makro Transportes
============================================
A unidade é a ATIVIDADE. A aderência é diária e por atividade:

    aderência = atividades concluídas ÷ atividades programadas do plano

O que NÃO entra nessa conta, e por quê:
  · empresa terceirizada — é serviço de fora, não mede a oficina da Makro;
  · extra programação — não dá para cobrar cumprimento de uma coisa que
    nunca foi programada;
  · cancelada — sai dos dois lados da divisão.
Os três têm contador próprio, e ao lado da aderência do plano fica o
CUMPRIMENTO GERAL, que inclui o extra: é a diferença entre os dois números
que mostra o tamanho do imprevisto na semana.

Programar é escrever o NÚMERO DA SEMANA (35, 36…) e o DIA. A data sai
sozinha do calendário ISO do ano configurado.

    python3 montar_planilha.py dados.json [--ano 2026]

`dados.json` vem do ler_planilha.py (planilha em uso) ou do extrair_pdf.py.
"""
import sys, json, io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule, Rule
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                   ParagraphProperties, CharacterProperties)

SAIDA="/home/user/gestaomkt/planilha/Programacao_Servicos_Makro.xlsx"

NAVY="FF13164E"; NAVY2="FF2A3170"; RED="FFE4002A"
CINZA="FFD9DEE8"; CINZA_C="FFEEF1F7"; AMARELO="FFFFF3CF"
BORDA="FFC8D0DE"; BRANCO="FFFFFFFF"; TINTA="FF14181F"; T2="FF4A5464"
OK_V="FFE3F5EA"; OK_T="FF116B3E"; AL_V="FFFDF0D8"; AL_T="FF8A5A05"
RU_V="FFFDE4E7"; RU_T="FFB00020"; AZ_V="FFE7EEFB"; AZ_T="FF1749C4"
EX_V="FFFFE9CC"; EX_T="FF9A4C00"

F=lambda **k: Font(name="Arial", **k)
fina=Side(style="thin", color=BORDA)
box=Border(left=fina,right=fina,top=fina,bottom=fina)
def fill(c): return PatternFill("solid", fgColor=c)
DFMT='[$-416]dd/mm/yyyy;@'          # dd/mm/aaaa em qualquer Excel

PRIM=5; ULT=1004; LIN_LISTA=120
DIAS_PT=["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"]

# ═══════════════════════════════════ entrada
args=[a for a in sys.argv[1:] if not a.startswith("--")]
ANO=2026
if "--ano" in sys.argv: ANO=int(sys.argv[sys.argv.index("--ano")+1])
# --puxar 35:36 leva o que não saiu na semana 35 para a 36. É a "continuidade"
# de toda segunda-feira, e por isso é parâmetro e não regra fixa.
PUXAR=None
if "--puxar" in sys.argv:
    _de,_para=sys.argv[sys.argv.index("--puxar")+1].split(":")
    PUXAR=(int(_de),int(_para))
FONTE=json.load(io.open(args[0],encoding="utf-8")) if args else {"linhas":[],"listas":{}}

def d2(s):
    if not s: return None
    a,m,d=s.split("-"); return date(int(a),int(m),int(d))

def seg_iso(ano,n):
    """Segunda-feira da semana ISO n do ano."""
    q=date(ano,1,4)
    return q - timedelta(days=q.weekday()) + timedelta(weeks=n-1)

# ── estimativa inicial de dias ──
# Chute de partida na faixa que a oficina pratica: a maioria em 2 ou 3 dias e
# uma minoria em 4, dando média perto de 2,7. O sorteio é determinístico
# (crc32 do texto), então reconstruir a planilha não embaralha as estimativas.
# É ponto de partida para ajustar linha a linha, não medição.
import zlib
def _dias_estimados(frota,servico,atividade):
    h=zlib.crc32((frota+"|"+servico+"|"+atividade).encode("utf-8"))%20
    return 2 if h<9 else (3 if h<17 else 4)

# A coluna Marcar só serve para FORÇAR o que a data não sabe dizer: cancelada,
# concluída, ou ainda em programação. "Programada", "Em execução" e "Em
# andamento" a Situação já calcula sozinha pela janela de execução — deixá-las
# escritas ali é ruído que não muda nada. Saem na importação.
INERTE={"em execução","em execucao","em andamento","em curso","programada"}
def _marcar(v):
    return "" if str(v or "").strip().lower() in INERTE else (v or "")

def _num(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None

LINHAS=[]
for x in FONTE["linhas"]:
    d=d2(x.get("data")); o=d2(x.get("orig"))
    # A semana e o dia vêm prontos de uma planilha em uso. Só se derivam da
    # data quando a origem é o PDF, ou uma planilha do layout antigo.
    _sem=_num(x.get("semana")); _sor=_num(x.get("semorig")); _dp=_num(x.get("dias"))
    LINHAS.append(dict(
        os=x.get("os",""), frota=x.get("frota",""), servico=x.get("servico",""),
        atividade=x.get("atividade",""), tipo=x.get("tipo") or "Corretiva",
        origem=x.get("origem") or "Programada", equipe=(x.get("equipe") or [])[:3],
        semana=_sem if _sem else (d.isocalendar()[1] if d else None),
        dia=x.get("dia") or (DIAS_PT[d.weekday()] if d else ""),
        # estimativa que ele já ajustou na mão manda; o chute só entra quando
        # a atividade ainda não tem nenhuma
        dias=(_dp if _dp else
              (_dias_estimados(x.get("frota",""),x.get("servico",""),x.get("atividade",""))
               if x.get("atividade") else "")),
        concluida=d2(x.get("concluida")), marcar=_marcar(x.get("marcar","")),
        semorig=_sor if _sor else (o.isocalendar()[1] if o else None),
        motivo=x.get("motivo",""), obs=x.get("obs","")))

# Semana orig. igual à Semana não é reprogramação nenhuma: a coluna Reprog. já
# exige que sejam diferentes. Escrita, ela só faz parecer que tudo foi
# empurrado. Limpa.
for l in LINHAS:
    if l["semorig"] and l["semorig"]==l["semana"]: l["semorig"]=None

# ── continuidade: o que não saiu numa semana passa para a seguinte ──
# Ela entra no PLANO da semana nova, mesmo tendo chegado como extra: quando se
# programa uma atividade para a semana que vem, ela passa a ser cobrada. A
# semana de origem fica registrada, e é contra ela que sai a aderência ao plano
# original — reprogramar não limpa a ficha.
if PUXAR:
    _de,_para=PUXAR; _n=0
    for l in LINHAS:
        if (l["semana"]==_de and l["atividade"] and not l["concluida"]
                and l["marcar"]!="Cancelada"):
            if not l["semorig"]: l["semorig"]=_de
            l["semana"]=_para; l["origem"]="Programada"; _n+=1
    print(f"  continuidade: {_n} atividades da semana {_de} passaram para a {_para}")

# ── frotas que chegam do Pará em 31/08, segunda-feira da semana 36 ──
# Vieram anotadas em conjunto — F425/1038/1039, F621/433, F817/150, F818/745 —
# e foram cadastradas unidade a unidade, que é como o resto da lista é escrito.
# O 150 do conjunto F817/150 já estava lá e não se repete.
FROTAS_PARA=["F-425","F-1038","F-1039","F-621","F-433","F-817","F-818","F-745"]
SEM_CHEGADA=36
# Uma linha reservada por unidade, com frota, semana e dia prontos e o serviço
# em branco para ele escrever. Enquanto a Atividade estiver vazia a linha não
# entra em conta nenhuma: aderência, extra, vencidas e diária todas exigem
# atividade. Ela existe só para a frota já aparecer na grade da semana 36.
_com_linha={l["frota"] for l in LINHAS if l["frota"] and l["semana"]==SEM_CHEGADA}
for _f in FROTAS_PARA:
    if _f in _com_linha: continue
    LINHAS.append(dict(os="", frota=_f, servico="", atividade="", tipo="",
                       origem="Programada", equipe=[], semana=SEM_CHEGADA,
                       dia="Seg", dias="", concluida=None, marcar="",
                       semorig=None, motivo="", obs="Chegou do Pará em 31/08"))

LST=FONTE.get("listas",{})
def lista(nome,padrao):
    v=[x for x in LST.get(nome,[]) if x]
    return v or padrao
EXEC_LISTA=lista("Executantes",[])
FROTA_LISTA=sorted(set(lista("Frotas",[])) | set(FROTAS_PARA)
                   | {l["frota"] for l in LINHAS if l["frota"]})
TIPO_LISTA=lista("Tipo de serviço",["Corretiva","Preventiva","Preditiva","Melhoria","Inspeção"])
for l in LINHAS:
    for e in l["equipe"]:
        if e and e not in EXEC_LISTA: EXEC_LISTA.append(e)
EXEC_LISTA=sorted(set(EXEC_LISTA), key=lambda x:(("(externo)" in x), x.lower()))
MOTIVOS=lista("Motivo da reprogramação",[])
for m in ["Peça não chegou","Frota em viagem","Falta de mão de obra","Serviço maior que o previsto",
          "Mudou a prioridade","Oficina cheia","Aguardando terceiro","Box bloqueado",
          "Frota não chegou na oficina","Faltou ferramenta","Serviço extra entrou na frente"]:
    if m not in MOTIVOS: MOTIVOS.append(m)

wb=Workbook()

# ═══════════════════════════════════ LISTAS
ls=wb.active; ls.title="Listas"
COLS_L=[("Executantes",EXEC_LISTA),("Frotas",FROTA_LISTA),("Tipo de serviço",TIPO_LISTA),
        ("Marcar",["Programada","Em programação","Concluída","Cancelada"]),
        ("Motivo do atraso ou da mudança",MOTIVOS),("Dia",DIAS_PT)]
ls["A1"]="LISTAS DE APOIO"
ls["A1"].font=F(bold=True,size=11,color=BRANCO); ls["A1"].fill=fill(NAVY)
ls["A1"].alignment=Alignment(vertical="center",indent=1)
ls.merge_cells("A1:F1"); ls.row_dimensions[1].height=24
ls["A2"]=("Alimentam as caixas de seleção. Acrescente nomes aqui e as caixas acompanham. "
          "Frota nova você pode digitar direto na Programação — a planilha só avisa, não trava.")
ls["A2"].font=F(size=9,italic=True,color=T2); ls.merge_cells("A2:F2")
for j,(tit,vals) in enumerate(COLS_L,start=1):
    c=ls.cell(row=4,column=j,value=tit)
    c.font=F(bold=True,size=10,color=BRANCO); c.fill=fill(NAVY2)
    c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=box
    for i in range(5,LIN_LISTA+1):
        cc=ls.cell(row=i,column=j)
        if i-5<len(vals): cc.value=vals[i-5]
        cc.font=F(size=10); cc.border=box
ls.row_dimensions[4].height=26
for j,w in enumerate([26,12,16,14,30,8],start=1):
    ls.column_dimensions[get_column_letter(j)].width=w
ls.sheet_view.showGridLines=False; ls.freeze_panes="A5"
ls.page_setup.orientation="portrait"; ls.page_setup.fitToWidth=1
ls.page_setup.fitToHeight=1; ls.sheet_properties.pageSetUpPr.fitToPage=True
ls.print_area=f"A1:F{LIN_LISTA}"

# ═══════════════════════════════════ PROGRAMAÇÃO
pg=wb.create_sheet("Programação")
COLS=[("A","Nº",5,"c"),("B","Situação",19,"c"),
      ("C","OS",11,"1"),("D","Frota",9,"1"),("E","Serviço",25,"1"),
      ("F","Atividade",40,"1"),("G","Tipo",11,"1"),("H","Origem",12,"1"),
      ("I","Executante 1",16,"q"),("J","Executante 2",16,"q"),("K","Executante 3",16,"q"),
      ("L","Semana",8,"2"),("M","Dia",7,"2"),("N","Dias prev.",9,"2"),
      ("O","Início",11,"c"),("P","Fim do serviço",13,"c"),
      ("Q","Concluída em",12,"3"),("R","Marcar",12,"3"),("S","Semana orig.",11,"3"),
      ("T","Motivo do atraso / da mudança",26,"3"),("U","Obs.",22,"3"),
      ("V","Oficina",12,"c"),("W","Plano",11,"c"),("X","No prazo",8,"c"),
      ("Y","No plano",8,"c"),("Z","Atraso",8,"c"),("AA","Reprog.",8,"c"),("AB","Concl.",8,"c"),
      ("AC","Venc?",7,"c"),("AD","Fração",8,"c"),
      ("AE","Ordem dia",9,"c"),("AF","Ordem atr.",9,"c"),("AG","Quem",22,"c"),
      ("AH","Entra dia",9,"c"),("AI","Entra atr.",9,"c")]
CORB={"1":AMARELO,"q":"FFE9F3E6","2":"FFDCE9FA","3":"FFEDEFF4","c":CINZA}

pg["A1"]="PROGRAMAÇÃO DE SERVIÇOS  ·  MAKRO TRANSPORTES  ·  uma linha por atividade"
pg["A1"].font=F(bold=True,size=13,color=BRANCO); pg["A1"].fill=fill(NAVY)
pg["A1"].alignment=Alignment(vertical="center",indent=1)
pg.merge_cells("A1:AI1"); pg.row_dimensions[1].height=30
for a,b,txt,g in [("A","B","CALCULADO","c"),("C","H","1 · O QUE É O SERVIÇO","1"),
                  ("I","K","2 · QUEM FAZ — até três","q"),
                  ("L","N","3 · PROGRAMAR — semana, dia e quantos dias leva","2"),
                  ("O","P","CALCULADO","c"),("Q","U","4 · SÓ QUANDO ACONTECER","3"),
                  ("V","AI","CALCULADO — não digite aqui","c")]:
    pg.merge_cells(f"{a}2:{b}2")
    c=pg[f"{a}2"]; c.value=txt
    c.font=F(bold=True,size=9,color=T2); c.fill=fill(CORB[g])
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=box
pg.row_dimensions[2].height=18
for letra,tit,larg,grp in COLS:
    c=pg[letra+"3"]; c.value=tit
    c.font=F(bold=True,size=9.5,color=BRANCO)
    c.fill=fill(NAVY if grp in "12q3" else NAVY2)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=box
    pg.column_dimensions[letra].width=larg
pg.row_dimensions[3].height=32

DICAS={
 "B":"Sai sozinha, e SÓ ESTA CÉLULA muda de cor — o resto da linha fica limpo.\n\n"
     "Em programação · Na carteira · Falta a semana · Programada · "
     "Em execução · Fecha hoje · VENCIDA · "
     "Concluída · Concluída com atraso · Cancelada",
 "D":"Pode digitar uma frota que ainda não está na lista — a planilha avisa mas "
     "deixa passar. Depois acrescente na aba Listas para ela virar opção.",
 "F":"A ATIVIDADE é a unidade: “trocar as lonas”, “regular a catraca”.\n\n"
     "Um serviço com três atividades ocupa três linhas. A aderência sai daqui.",
 "H":"PROGRAMADA: estava no plano da semana.\nEXTRA: entrou depois — quebra, "
     "urgência, pedido da operação.\n\nExtra fica FORA da aderência do plano e "
     "tem contador próprio. Ao lado dela, o CUMPRIMENTO GERAL inclui o extra: "
     "a diferença entre os dois é o tamanho do imprevisto.",
 "I":"Até três pessoas na mesma atividade. Ela entra na carga de cada uma e "
     "conta uma vez só no total.\n\nEmpresa externa: escreva com “(externo)” no "
     "fim — serviço de terceiro fica FORA da aderência interna.",
 "L":"O NÚMERO da semana no ano (35, 36…), não a data. O ano fica na aba Semana.\n\n"
     "Para repetir a semana passada: copie as linhas, cole no fim e troque só "
     "este número.",
 "N":"Quantos dias o SERVIÇO leva por inteiro. Serve para dimensionar a semana "
     "— é a soma disto que vira a diária.\n\nO atraso da atividade continua sendo "
     "medido pelo DIA dela: cada linha tem o seu dia, e não a folga do serviço todo.",
 "Q":"Preencher esta data já conclui a atividade.",
 "AD":"1 dividido por quantos executantes a atividade tem. É o que alimenta a "
      "coluna PESO da aba Semana.\n\nUma atividade com três pessoas conta 1 no "
      "total da semana e 1/3 para cada uma.",
 "S":"Só ao REPROGRAMAR: o número da semana em que ela estava antes.",
 "T":"Por que atrasou, ou por que mudou de semana. Escolha da lista ou escreva.",
}
for col,txt in DICAS.items(): pg[col+"3"].comment=Comment(txt,"PCM")

for i,d in enumerate(LINHAS):
    r=PRIM+i
    eq=(list(d["equipe"])+["","",""])[:3]
    for col,val in (("C",d["os"]),("D",d["frota"]),("E",d["servico"]),("F",d["atividade"]),
                    ("G",d["tipo"]),("H",d["origem"]),("I",eq[0]),("J",eq[1]),("K",eq[2]),
                    ("L",d["semana"]),("M",d["dia"]),("N",d["dias"]),
                    ("Q",d["concluida"]),("R",d["marcar"]),("S",d["semorig"]),
                    ("T",d["motivo"]),("U",d["obs"])):
        if val not in (None,""): pg[f"{col}{r}"]=val

ANOREF="Semana!$J$3"; EXPREF="Semana!$M$3"
# Duas datas que não se confundem:
#   HOJEREF  — a data que ele ESCOLHE olhar; move a aba Hoje e as listas.
#   HOJEREAL — o hoje de verdade; move a Situação. Se a Situação seguisse a
#              data escolhida, espiar a quinta reescreveria o status de mil
#              linhas.
HOJEREF="Hoje!$C$3"; HOJEREAL="Hoje!$L$1"; AGORA="Hoje!$L$2"
SEG=f'(DATE({ANOREF},1,4)-WEEKDAY(DATE({ANOREF},1,4),3))'   # segunda da semana 1
for r in range(PRIM,ULT+1):
    # Âncora na linha 4 pelo mesmo motivo das colunas de ordem: ancorada na 5,
    # apagar a primeira linha de dados levava a numeração toda junto.
    pg[f"A{r}"]=f'=IF($F{r}="","",COUNTA($F${PRIM-1}:$F{r}))'
    pg[f"O{r}"]=(f'=IF(OR($L{r}="",$M{r}=""),"",'
                 f'{SEG}+($L{r}-1)*7+MATCH($M{r},Listas!$F$5:$F$11,0)-1)')
    pg[f"P{r}"]=f'=IF($O{r}="","",$O{r}+MAX(1,IF($N{r}="",1,$N{r}))-1)'
    pg[f"W{r}"]=(f'=IF($O{r}="","",IF($S{r}="",$O{r},{SEG}+($S{r}-1)*7+'
                 f'IF($M{r}="",0,MATCH($M{r},Listas!$F$5:$F$11,0)-1)))')
    # Vencida é só depois que a janela de execução FECHOU. Uma atividade de
    # três dias que começa segunda está em execução até quarta — chamar isso de
    # vencida já na segunda era o que enchia a planilha de vermelho sem motivo.
    # No último dia ela só vence depois da hora de fechar a oficina.
    # Fica em coluna própria para a fórmula da Situação caber: inteira, ela
    # estourava e voltava #VALOR! em toda linha.
    # TODAY() e NOW() são voláteis: dentro das mil linhas, o Excel refazia
    # cinco mil chamadas a cada tecla. Elas passaram a ser calculadas uma vez
    # só, no bloco motor da aba Hoje, e aqui só se lê o resultado.
    pg[f"AC{r}"]=(f'=IF($P{r}="",0,IF($P{r}<{HOJEREAL},1,'
                  f'IF(AND($P{r}={HOJEREAL},{AGORA}>{EXPREF}),1,0)))')
    # A janela manda no status: antes dela, Programada; dentro, Em execução;
    # no último dia, Fecha hoje; depois, VENCIDA.
    # Linha com frota e sem atividade é a linha reservada: ela não conta em
    # lugar nenhum, e o status diz o que falta para ela passar a valer.
    # Dia escrito e semana em branco não gera Início, e sem Início a atividade
    # some: some da semana, da aba Hoje e de todo indicador. Ele tinha 37
    # assim, marcadas para segunda e invisíveis. Agora elas se anunciam.
    pg[f"B{r}"]=(f'=IF($F{r}="",IF($D{r}="","","Falta a atividade"),'
      f'IF($R{r}="Cancelada","Cancelada",'
      f'IF($AB{r}=1,IF(N($Z{r})>0,"Concluída com atraso","Concluída"),'
      f'IF($R{r}="Em programação","Em programação",'
      f'IF($O{r}="",IF($M{r}="","Na carteira","Falta a semana"),'
      f'IF($AC{r}=1,"VENCIDA",'
      f'IF($O{r}>{HOJEREAL},"Programada",'
      f'IF($P{r}>{HOJEREAL},"Em execução","Fecha hoje"))))))))')
    pg[f"V{r}"]=(f'=IF($F{r}="","",IF(ISNUMBER(SEARCH("(externo)",$I{r}&$J{r}&$K{r})),'
                 f'"Terceirizada","Interna"))')
    pg[f"X{r}"]=f'=IF($F{r}="","",IF(OR($Q{r}="",$O{r}=""),0,IF($Q{r}<=$O{r},1,0)))'
    pg[f"Y{r}"]=f'=IF($F{r}="","",IF(OR($Q{r}="",$W{r}=""),0,IF($Q{r}<=$W{r},1,0)))'
    pg[f"Z{r}"]=f'=IF(OR($Q{r}="",$O{r}=""),"",MAX(0,$Q{r}-$O{r}))'
    pg[f"AA{r}"]=f'=IF($F{r}="","",IF(AND($S{r}<>"",$L{r}<>"",$S{r}<>$L{r}),1,0))'
    pg[f"AB{r}"]=(f'=IF($F{r}="","",IF($R{r}="Cancelada",0,'
                  f'IF(OR($Q{r}<>"",$R{r}="Concluída"),1,0)))')
    # Quanto desta atividade cabe a CADA um que participou. Três executantes
    # numa atividade dão 1/3 para cada, e a soma volta a fechar com o número
    # de atividades — que é o que a linha de TOTAL sempre mostrou. Sem isto,
    # a coluna por pessoa somava o dobro do total e parecia erro de conta.
    pg[f"AD{r}"]=(f'=IF($F{r}="","",IF(COUNTA($I{r}:$K{r})=0,0,'
                  f'1/COUNTA($I{r}:$K{r})))')
    # ── numeração que alimenta a aba Hoje ──
    # NUNCA voltar a fazer isto em cadeia (cada linha lendo N($AE{r-1})).
    # Apagar uma linha no Excel fazia a linha seguinte apontar para uma linha
    # que não existe mais, e o #REF! descia por todas as outras: uma linha
    # apagada produziu 2.593 células quebradas.
    # A bandeira só olha a própria linha; a ordem é uma soma acumulada de uma
    # âncora fixa até aqui. Apagar linha só encurta o intervalo.
    pg[f"AH{r}"]=(f'=IF(AND($F{r}<>"",$O{r}<>"",$O{r}<={HOJEREF},'
                  f'$P{r}>={HOJEREF},$R{r}<>"Cancelada"),1,0)')
    pg[f"AI{r}"]=(f'=IF(AND($F{r}<>"",$P{r}<>"",$P{r}<{HOJEREF},'
                  f'$AB{r}=0,$R{r}<>"Cancelada"),1,0)')
    # A âncora é a linha 4, vazia e dentro do painel congelado. Ancorar na 5,
    # que é a primeira linha de dados, quebraria se ele apagasse justo ela.
    pg[f"AE{r}"]=f'=IF($AH{r}=0,"",SUM($AH$4:$AH{r}))'
    pg[f"AF{r}"]=f'=IF($AI{r}=0,"",SUM($AI$4:$AI{r}))'
    pg[f"AG{r}"]=(f'=IF($F{r}="","",$I{r}&IF($J{r}="",""," · "&$J{r})'
                  f'&IF($K{r}="",""," · "&$K{r}))')
    for letra,tit,larg,grp in COLS:
        c=pg[letra+str(r)]
        c.font=F(size=10, color=TINTA if grp in "12q3" else T2)
        c.border=box
        if grp=="c": c.fill=fill(CINZA_C)
        elif grp=="2": c.fill=fill("FFF4F9FF")
        elif grp=="q": c.fill=fill("FFF6FBF5")
        if letra in ("O","P","Q","W"): c.number_format=DFMT
        # O Excel come o zero à esquerda de uma OS como 021188 se a coluna
        # for geral. Texto explícito segura.
        if letra=="C": c.number_format="@"
        if letra in ("A","B","G","H","L","M","N","R","S","V","X","Y","Z","AA","AB","AC","AD","AE","AF","AH","AI"):
            c.alignment=Alignment(horizontal="center")
        if letra in ("L","N","S"): c.number_format="0"
    pg.row_dimensions[r].height=16.5

# ── caixas de seleção ──
def dv(col, formula, titulo, msg, brando=False):
    v=DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    v.errorTitle=titulo; v.error=msg; v.showErrorMessage=True
    if brando: v.errorStyle="warning"
    pg.add_data_validation(v); v.add(f"{col}{PRIM}:{col}{ULT}")
# Frota nova não pode travar o trabalho: a caixa avisa e deixa passar.
dv("D", f"=Listas!$B$5:$B${LIN_LISTA}", "Frota nova",
   "Esta frota ainda não está na aba Listas.\n\nPode continuar — depois "
   "acrescente o nome lá para ela virar opção da caixa.", brando=True)
dv("G", f"=Listas!$C$5:$C${LIN_LISTA}", "Tipo", "Escolha um tipo da aba Listas.")
dv("H", '"Programada,Extra"', "Origem", "Programada (estava no plano) ou Extra (entrou depois).")
for cx in ("I","J","K"):
    dv(cx, f"=Listas!$A$5:$A${LIN_LISTA}", "Executante",
       "Nome fora da lista.\n\nPode continuar — depois acrescente na aba Listas.", brando=True)
dv("M", "=Listas!$F$5:$F$11", "Dia", "Seg, Ter, Qua, Qui, Sex, Sáb ou Dom.")
dv("R", "=Listas!$D$5:$D$8", "Marcar",
   "Deixe vazio (a situação sai sozinha), ou escolha: Programada / "
   "Em programação / Concluída / Cancelada.")
dv("T", f"=Listas!$E$5:$E${LIN_LISTA}", "Motivo", "Escolha da aba Listas ou escreva o seu.", brando=True)
vn=DataValidation(type="whole", operator="between", formula1="1", formula2="53", allow_blank=True)
vn.errorTitle="Semana"; vn.error="O número da semana vai de 1 a 53."; vn.showErrorMessage=True
pg.add_data_validation(vn); vn.add(f"L{PRIM}:L{ULT}"); vn.add(f"S{PRIM}:S{ULT}")
vd=DataValidation(type="whole", operator="between", formula1="1", formula2="60", allow_blank=True)
vd.errorTitle="Dias previstos"; vd.error="De 1 a 60 dias."; vd.showErrorMessage=True
pg.add_data_validation(vd); vd.add(f"N{PRIM}:N{ULT}")

# ── a cor fica na CÉLULA, não na linha inteira ──
# Pintar a linha toda virava borrão: com fundo colorido de ponta a ponta não
# se lia mais a atividade nem a observação. Agora só muda de cor a célula que
# carrega aquela informação — o texto fica no branco.
FAIXA=f"A{PRIM}:AI{ULT}"
PINTA=[("VENCIDA",             RU_V,       RU_T,       True),
       ("Fecha hoje",          EX_V,       EX_T,       True),
       ("Em execução",         AZ_V,       AZ_T,       True),
       ("Falta a atividade",   "FFF6E8CE", "FF8A5A05", False),
       ("Falta a semana",      "FFF6E8CE", "FF8A5A05", True),
       ("Concluída",           OK_V,       OK_T,       True),
       ("Concluída com atraso",AL_V,       AL_T,       True),
       ("Em programação",      AZ_V,       AZ_T,       True),
       ("Programada",          "FFEDF2FC", "FF2C4272",  False),
       ("Na carteira",         "FFF2F4F8", T2,          False),
       ("Cancelada",           "FFE6E9EF", "FF7C8698",  False)]
for txt,bg,fg,negrito in PINTA:
    rg_=Rule(type="expression", stopIfTrue=True,
             dxf=DifferentialStyle(fill=PatternFill(bgColor=bg),
                                   font=Font(color=fg,bold=negrito)))
    rg_.formula=[f'$B{PRIM}="{txt}"']
    pg.conditional_formatting.add(f"B{PRIM}:B{ULT}", rg_)

# cancelada: sem fundo, só o texto riscado e apagado no que descreve o serviço
canc=Rule(type="expression", dxf=DifferentialStyle(
    font=Font(strike=True, color="FF98A1B2")))
canc.formula=[f'$B{PRIM}="Cancelada"']
pg.conditional_formatting.add(f"C{PRIM}:H{ULT}", canc)

# ── EXTRA PROGRAMAÇÃO: a marca mais forte da planilha ──
# Fundo laranja cheio, texto escuro em negrito e moldura grossa em volta da
# célula. O número da linha ganha o mesmo laranja, para o extra se achar de
# longe sem precisar ler a coluna Origem.
EXV="FFFFC97A"; EXT="FF6E2E00"
gro=Side(style="medium", color=EXT)
exh=Rule(type="expression", stopIfTrue=True, dxf=DifferentialStyle(
    fill=PatternFill(bgColor=EXV), font=Font(color=EXT,bold=True),
    border=Border(left=gro,right=gro,top=gro,bottom=gro)))
exh.formula=[f'$H{PRIM}="Extra"']
pg.conditional_formatting.add(f"H{PRIM}:H{ULT}", exh)
exn=Rule(type="expression", dxf=DifferentialStyle(
    fill=PatternFill(bgColor=EXV), font=Font(color=EXT,bold=True)))
exn.formula=[f'$H{PRIM}="Extra"']
pg.conditional_formatting.add(f"A{PRIM}:A{ULT}", exn)

# ── as datas avisam sozinhas, só com a cor da letra ──
for txt,cor in (("VENCIDA",RU_T),("Fecha hoje",EX_T),("Em execução",AZ_T)):
    rd=Rule(type="expression", stopIfTrue=True,
            dxf=DifferentialStyle(font=Font(color=cor,bold=True)))
    rd.formula=[f'$B{PRIM}="{txt}"']
    pg.conditional_formatting.add(f"O{PRIM}:P{ULT}", rd)
# a data de conclusão fica verde no instante em que é escrita
rq=Rule(type="expression", dxf=DifferentialStyle(
    fill=PatternFill(bgColor=OK_V), font=Font(color=OK_T,bold=True)))
rq.formula=[f'AND($Q{PRIM}<>"",$F{PRIM}<>"")']
pg.conditional_formatting.add(f"Q{PRIM}:Q{ULT}", rq)
# motivo escrito: amarelo claro, para saber de relance quem já foi justificado
rt=Rule(type="expression", dxf=DifferentialStyle(fill=PatternFill(bgColor="FFFFF6DC")))
rt.formula=[f'$T{PRIM}<>""']
pg.conditional_formatting.add(f"T{PRIM}:T{ULT}", rt)
# régua entre um serviço e outro
reg=Rule(type="expression", dxf=DifferentialStyle(
    border=Border(top=Side(style="medium", color="FF8494B0"))))
reg.formula=[f'AND($F{PRIM}<>"",$D{PRIM}&$E{PRIM}<>$D{PRIM-1}&$E{PRIM-1})']
pg.conditional_formatting.add(FAIXA, reg)

pg.freeze_panes="G5"
pg.auto_filter.ref=f"A3:AI{ULT}"
pg.sheet_view.showGridLines=False

# ═══════════════════════════════════ referências
PROG="Programação"
def rg(c): return f"{PROG}!${c}${PRIM}:${c}${ULT}"
SIT=rg("B"); ATIV=rg("F"); FROTA_=rg("D"); TIPO_=rg("G"); ORIG=rg("H")
EQ=(rg("I"),rg("J"),rg("K")); SEM=rg("L"); DIAS=rg("N")
INI=rg("O"); FIM=rg("P"); CONCLEM=rg("Q"); MARC=rg("R")
OFIC=rg("V"); PLANOD=rg("W"); NOPRAZO=rg("X"); NOPLANO=rg("Y")
ATRASO=rg("Z"); REP=rg("AA"); CONCL=rg("AB"); FRACAO=rg("AD")
ORD_DIA=rg("AE"); ORD_ATR=rg("AF"); QUEM=rg("AG"); SEMORIG=rg("S")
INT=f'{OFIC},"Interna"'          # só a oficina da Makro
NC =f'{MARC},"<>Cancelada"'
NEX=f'{ORIG},"<>Extra"'
BASE=f'{NC},{INT}'               # tudo que a oficina interna tinha para fazer
PLAN=f'{BASE},{NEX}'             # e que estava no plano

def porpessoa(alvo, extra, agreg="COUNTIFS", faixa=None):
    ini=f"SUMIFS({faixa}," if agreg=="SUMIFS" else "COUNTIFS("
    return "+".join(f'{ini}{c},{alvo},{extra})' for c in EQ)
def titulo(ws, texto, ate, altura=30, tam=13):
    ws["A1"]=texto
    ws["A1"].font=F(bold=True,size=tam,color=BRANCO); ws["A1"].fill=fill(NAVY)
    ws["A1"].alignment=Alignment(vertical="center",indent=1)
    ws.merge_cells(f"A1:{ate}1"); ws.row_dimensions[1].height=altura
def cabec(ws, linha, itens, cor=NAVY2):
    for col,txt in itens:
        c=ws[f"{col}{linha}"]; c.value=txt
        c.font=F(bold=True,size=9.5,color=BRANCO); c.fill=fill(cor)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=box

# ═══════════════════════════════════ SEMANA
sm=wb.create_sheet("Semana")
titulo(sm,"GRADE DA SEMANA  ·  atividades por executante e por dia","N")
sm["A3"]="SEMANA Nº"
sm["A3"].font=F(bold=True,size=12,color=NAVY)
sm["A3"].alignment=Alignment(vertical="center",indent=1); sm.merge_cells("A3:B3")
# Abre na semana mais adiantada que tem programação: é a que está sendo tocada.
# Abrir na mais antiga fazia a planilha começar num retrato já encerrado.
sem_ini=max([l["semana"] for l in LINHAS if l["semana"]] or [date.today().isocalendar()[1]])
sm["C3"]=sem_ini
c=sm["C3"]; c.font=F(bold=True,size=18,color=NAVY); c.fill=fill(AMARELO)
c.border=Border(*[Side(style="medium",color=NAVY)]*4)
c.alignment=Alignment(horizontal="center",vertical="center"); c.number_format="0"
c.comment=Comment("O número da semana no ano — 35, 36, 40…\n\n"
 "Troque aqui e a grade inteira acompanha: os dias no topo, os números de "
 "cada executante e o resumo.","PCM")
vsem=DataValidation(type="whole",operator="between",formula1="1",formula2="53",allow_blank=False)
vsem.errorTitle="Semana"; vsem.error="De 1 a 53."; vsem.showErrorMessage=True
sm.add_data_validation(vsem); vsem.add("C3")
sm["D3"]="de"; sm["F3"]="a"
sm["E3"]=f'={SEG}+($C$3-1)*7'
sm["G3"]="=$E$3+6"
for ref in ("E3","G3"):
    x=sm[ref]; x.font=F(bold=True,size=12,color=T2); x.number_format=DFMT
    x.alignment=Alignment(horizontal="center",vertical="center")
    x.fill=fill(CINZA_C); x.border=box
for ref in ("D3","F3"):
    sm[ref].font=F(size=10,color=T2)
    sm[ref].alignment=Alignment(horizontal="center",vertical="center")
sm["I3"]="Ano:"; sm["I3"].font=F(bold=True,size=10)
sm["I3"].alignment=Alignment(horizontal="right",vertical="center")
sm["J3"]=ANO
x=sm["J3"]; x.font=F(bold=True,size=12,color=NAVY); x.fill=fill(AMARELO)
x.border=box; x.alignment=Alignment(horizontal="center",vertical="center"); x.number_format="0"
x.comment=Comment("Ano do calendário. É dele e do número da semana que saem "
 "as datas de todas as atividades.","PCM")
sm["K3"]="Expediente até:"; sm.merge_cells("K3:L3")
sm["K3"].font=F(bold=True,size=10)
sm["K3"].alignment=Alignment(horizontal="right",vertical="center")
sm["M3"]=0.75
x=sm["M3"]; x.font=F(bold=True,size=12,color=NAVY); x.fill=fill(AMARELO)
x.border=box; x.alignment=Alignment(horizontal="center",vertical="center")
x.number_format="hh:mm"
x.comment=Comment("Hora em que a oficina fecha.\n\nPassado este horário, o que "
 "vencia hoje e não foi concluído passa a contar como VENCIDA — não fica "
 "esperando o dia seguinte para aparecer.","PCM")
sm.row_dimensions[3].height=30

jan=f'{INI},">="&$E$3,{INI},"<="&$G$3'
sm["A4"]="NESTA SEMANA  ·  só a oficina interna; serviço de terceiro fica de fora da conta"
sm["A4"].font=F(bold=True,size=10,color=NAVY)
sm["A4"].alignment=Alignment(vertical="center",indent=1)
sm.merge_cells("A4:N4"); sm.row_dimensions[4].height=20
RES=[("A","C","PROGRAMADAS",f'=COUNTIFS({jan},{PLAN})',"0",NAVY,
      "atividades do plano da semana, sem extra, sem terceiro e sem cancelada"),
     ("D","E","CONCLUÍDAS",f'=SUMIFS({CONCL},{jan},{NEX},{INT})',"0",OK_T,
      "do plano, quantas saíram"),
     ("F","G","ADERÊNCIA",'=IF($A$6=0,"",$D$6/$A$6)',"0%",NAVY,
      "concluídas ÷ programadas — mede se o plano foi cumprido"),
     ("H","I","CUMPRIMENTO GERAL",f'=IFERROR(SUMIFS({CONCL},{jan},{INT})/COUNTIFS({jan},{BASE}),"")',"0%",AZ_T,
      "tudo que a oficina fez ÷ tudo que ela teve, incluindo o extra. "
      "A diferença para a aderência é o tamanho do imprevisto."),
     ("J","K","EXTRA",f'=COUNTIFS({jan},{ORIG},"Extra",{BASE})',"0",EX_T,
      "entraram fora da programação"),
     ("L","M","VENCIDAS",f'=COUNTIFS({SIT},"VENCIDA",{jan},{INT})',"0",RU_T,
      "o dia passou e a atividade não foi concluída")]
for c1,c2,rot,fml,fmt,cor,dica in RES:
    sm.merge_cells(f"{c1}5:{c2}5"); sm.merge_cells(f"{c1}6:{c2}6")
    a=sm[f"{c1}5"]; a.value=rot
    a.font=F(bold=True,size=8.5,color=T2); a.alignment=Alignment(horizontal="center",wrap_text=True)
    a.fill=fill(CINZA_C); a.comment=Comment(dica,"PCM")
    b=sm[f"{c1}6"]; b.value=fml
    b.font=F(bold=True,size=17,color=cor); b.number_format=fmt
    b.alignment=Alignment(horizontal="center",vertical="center")
    b.fill=fill(EX_V if rot=="EXTRA" else CINZA_C)
    for cc in (c1,c2):
        sm[f"{cc}5"].border=box; sm[f"{cc}6"].border=box
sm.row_dimensions[5].height=22; sm.row_dimensions[6].height=28
sm["A7"]=(f'="Fora da conta interna:  "&COUNTIFS({jan},{OFIC},"Terceirizada",{NC})&'
          f'" em empresa terceirizada  ·  "&COUNTIFS({jan},{MARC},"Cancelada")&" cancelada(s)"&'
          f'"      |      Na carteira, sem dia:  "&COUNTIFS({INI},"",{ATIV},"<>",{NC})&" atividade(s)"')
sm["A7"].font=F(size=10,italic=True,color=T2)
sm["A7"].alignment=Alignment(vertical="center",indent=1); sm.merge_cells("A7:N7")
sm.row_dimensions[7].height=18

sm["A9"]="ONDE TEM PROGRAMAÇÃO"
sm["A9"].font=F(bold=True,size=10,color=NAVY)
sm["A9"].alignment=Alignment(vertical="center",indent=1); sm.merge_cells("A9:C9")
sm["D9"]="atividades em cada semana — a do meio é a que você escolheu"
sm["D9"].font=F(size=9,italic=True,color=T2); sm.merge_cells("D9:M9")
NAV=["E","F","G","H","I","J","K","L","M"]
for k,col in enumerate(NAV):
    off=k-4
    a=sm[f"{col}10"]; a.value=f"=$C$3+{off}" if off else "=$C$3"
    a.number_format='"sem "0'
    a.font=F(bold=True,size=10,color=BRANCO if off==0 else T2)
    a.fill=fill(NAVY if off==0 else CINZA_C)
    a.alignment=Alignment(horizontal="center"); a.border=box
    b=sm[f"{col}11"]
    b.value=(f'=COUNTIFS({SEM},{col}$10,{ATIV},"<>",{NC})')
    b.font=F(bold=True,size=12,color=NAVY if off==0 else T2)
    b.alignment=Alignment(horizontal="center"); b.border=box
    b.fill=fill(AZ_V if off==0 else BRANCO)
sm["D10"]="semana"; sm["D11"]="atividades"
for r in (10,11):
    sm[f"D{r}"].font=F(size=9,italic=True,color=T2)
    sm[f"D{r}"].alignment=Alignment(horizontal="right",vertical="center")
sm.conditional_formatting.add("E11:M11", CellIsRule(operator="equal", formula=["0"],
    font=F(size=12,color="FFB9C2D0")))
sm.row_dimensions[11].height=20

# ── a grade ──
LCAB=13
DS=["B","C","D","E","F","G","H"]
cabec(sm,LCAB,[("A","Executante")]+[(c,"") for c in DS]+
      [("I","Atividades"),("J","Dias prev."),("K","Concluídas"),("L","% feito"),
       ("M","Vencidas"),("N","Peso")])
for i,col in enumerate(DS):
    c=sm[f"{col}{LCAB}"]; c.value=f"=$E$3+{i}" if i else "=$E$3"
    c.number_format='[$-416]ddd\\ dd/mm;@'
sm[f"J{LCAB}"].comment=Comment("A diária: soma dos dias previstos das atividades da "
 "pessoa na semana. Uma atividade de 3 dias pesa 3 aqui e 1 na coluna Total.","PCM")
sm[f"I{LCAB}"].comment=Comment("Tudo que a pessoa tem na semana — plano e extra juntos. "
 "A grade é carga de trabalho; quem separa plano de extra é a faixa lá em cima.\n\n"
 "ATENÇÃO: atividade dividida conta CHEIA para cada um que participou. Por isso "
 "esta coluna soma mais que o total da semana — quem fecha com o total é o PESO.","PCM")
sm[f"N{LCAB}"].comment=Comment("A atividade dividida entre quem participou: três "
 "executantes dão 1/3 para cada.\n\nÉ esta coluna que FECHA com o total da semana. "
 "Serve para dizer quanto da carga foi de cada um sem contar a mesma atividade "
 "duas vezes.","PCM")
sm.row_dimensions[LCAB].height=30

NEX_G=max(16,len(EXEC_LISTA)+4); P0=LCAB+1; P1=P0+NEX_G-1
LSEM=P1+1; LTOT=P1+2
for k in range(NEX_G):
    r=P0+k
    sm[f"A{r}"]=f'=IF(Listas!A{5+k}="","",Listas!A{5+k})'
    for i,col in enumerate(DS):
        sm[f"{col}{r}"]=f'=IF($A{r}="","",{porpessoa(f"$A{r}", f"{INI},{col}${LCAB},{NC}")})'
    sm[f"I{r}"]=f'=IF($A{r}="","",SUM($B{r}:$H{r}))'
    sm[f"J{r}"]=f'=IF($A{r}="","",{porpessoa(f"$A{r}", f"{jan},{NC}", "SUMIFS", DIAS)})'
    sm[f"K{r}"]=f'=IF($A{r}="","",{porpessoa(f"$A{r}", jan, "SUMIFS", CONCL)})'
    sm[f"L{r}"]=f'=IF(OR($A{r}="",$I{r}=0),"",$K{r}/$I{r})'
    _venc=SIT+',"VENCIDA",'+jan
    sm[f"M{r}"]=f'=IF($A{r}="","",{porpessoa(f"$A{r}", _venc)})'
    sm[f"N{r}"]=f'=IF($A{r}="","",{porpessoa(f"$A{r}", f"{jan},{NC}", "SUMIFS", FRACAO)})'
sm[f"A{LSEM}"]="— sem executante definido —"
sm[f"A{LSEM}"].font=F(size=10,italic=True,color=T2)
for i,col in enumerate(DS):
    sm[f"{col}{LSEM}"]=f'=COUNTIFS({EQ[0]},"",{ATIV},"<>",{INI},{col}${LCAB},{NC})'
sm[f"I{LSEM}"]=f'=SUM($B{LSEM}:$H{LSEM})'
sm[f"J{LSEM}"]=f'=SUMIFS({DIAS},{EQ[0]},"",{jan},{NC})'
sm[f"K{LSEM}"]=f'=SUMIFS({CONCL},{EQ[0]},"",{jan})'
sm[f"L{LSEM}"]=f'=IF($I{LSEM}=0,"",$K{LSEM}/$I{LSEM})'
sm[f"M{LSEM}"]=f'=COUNTIFS({EQ[0]},"",{SIT},"VENCIDA",{jan})'
# atividade sem ninguém também tem peso, senão a coluna não fecha com o total
sm[f"N{LSEM}"]=f'=COUNTIFS({EQ[0]},"",{ATIV},"<>",{jan},{NC})'
sm[f"A{LTOT}"]="TOTAL — atividades distintas"
sm[f"A{LTOT}"].comment=Comment("Atividades DISTINTAS da semana. Não é a soma da "
 "coluna de cima: atividade dividida conta cheia para cada um que participou, "
 "então aquela coluna soma mais.\n\nQuem fecha com este total é o PESO.","PCM")
for i,col in enumerate(DS):
    sm[f"{col}{LTOT}"]=f'=COUNTIFS({ATIV},"<>",{INI},{col}${LCAB},{NC})'
sm[f"I{LTOT}"]=f'=COUNTIFS({jan},{ATIV},"<>",{NC})'
sm[f"J{LTOT}"]=f'=SUMIFS({DIAS},{jan},{NC})'
sm[f"K{LTOT}"]=f'=SUMIFS({CONCL},{jan})'
sm[f"L{LTOT}"]=f'=IF($I{LTOT}=0,"",$K{LTOT}/$I{LTOT})'
sm[f"M{LTOT}"]=f'=COUNTIFS({SIT},"VENCIDA",{jan})'
sm[f"N{LTOT}"]=f'=SUM($N{P0}:$N{LSEM})'
for r in range(P0,LTOT+1):
    for col in ["A"]+DS+["I","J","K","L","M","N"]:
        c=sm[f"{col}{r}"]; c.border=box; c.font=F(size=10)
        if col!="A": c.alignment=Alignment(horizontal="center")
        if col=="L": c.number_format="0%"
        if col=="N": c.number_format="0.0"
        if r==LTOT: c.font=F(size=10,bold=True,color=BRANCO); c.fill=fill(NAVY)
        elif r==LSEM: c.fill=fill(CINZA_C)
        elif col in ("I","J","K","L","M","N"): c.fill=fill(CINZA_C)
    sm.row_dimensions[r].height=17
sm.column_dimensions["A"].width=26
for col in DS: sm.column_dimensions[col].width=10
for col,w in (("I",11),("J",10),("K",11),("L",9),("M",10),("N",8)):
    sm.column_dimensions[col].width=w
sm.conditional_formatting.add(f"B{P0}:H{P1}",
    DataBarRule(start_type="num",start_value=0,end_type="num",end_value=8,
                color="FF9DB4DE",showValue=True))
sm.conditional_formatting.add(f"B{P0}:H{P1}", CellIsRule(operator="equal",
    formula=["0"], font=F(size=10,color="FFC3CBD8")))
sm.conditional_formatting.add(f"L{P0}:L{P1}", CellIsRule(operator="greaterThanOrEqual",
    formula=["0.85"], fill=fill(OK_V), font=F(bold=True,size=10,color=OK_T)))
sm.conditional_formatting.add(f"L{P0}:L{P1}", CellIsRule(operator="lessThan",
    formula=["0.6"], fill=fill(RU_V), font=F(bold=True,size=10,color=RU_T)))
sm.conditional_formatting.add(f"M{P0}:M{LSEM}", CellIsRule(operator="greaterThan",
    formula=["0"], fill=fill(RU_V), font=F(bold=True,size=10,color=RU_T)))
# ── a semana dia a dia ──
# O denominador é o TÉRMINO previsto, não o início: o que interessa saber é o
# que tinha de estar pronto naquele dia. É a aderência diária, e é ela que diz
# em que dia a semana descarrilou.
LDIA=LTOT+2
sm[f"A{LDIA}"]="ADERÊNCIA DIA A DIA  ·  o que tinha de fechar em cada dia"
sm[f"A{LDIA}"].font=F(bold=True,size=10,color=NAVY)
sm[f"A{LDIA}"].alignment=Alignment(vertical="center",indent=1)
sm.merge_cells(f"A{LDIA}:N{LDIA}"); sm.row_dimensions[LDIA].height=20
DIAG=[("Para fechar", lambda c: f'=COUNTIFS({FIM},{c}${LCAB},{ATIV},"<>",{BASE})',"0"),
      ("Fechou",      lambda c: f'=SUMIFS({CONCL},{FIM},{c}${LCAB},{BASE})',"0"),
      ("% do dia",    None,"0%")]
for i,(rot,f_,fmt) in enumerate(DIAG):
    r=LDIA+1+i
    sm[f"A{r}"]=rot
    sm[f"A{r}"].font=F(size=10,bold=(rot=="% do dia"))
    sm[f"A{r}"].alignment=Alignment(vertical="center",indent=1)
    for col in DS:
        sm[f"{col}{r}"]=(f'=IF({col}{LDIA+1}=0,"",{col}{LDIA+2}/{col}{LDIA+1})'
                         if f_ is None else f_(col))
        sm[f"{col}{r}"].number_format=fmt
    sm[f"I{r}"]=(f'=IF($I{LDIA+1}=0,"",$I{LDIA+2}/$I{LDIA+1})' if f_ is None
                 else f'=SUM($B{r}:$H{r})')
    sm[f"I{r}"].number_format=fmt
    for col in ["A"]+DS+["I"]:
        c=sm[f"{col}{r}"]; c.border=box
        if col!="A":
            c.font=F(size=10,bold=(rot=="% do dia")); c.fill=fill(CINZA_C)
            c.alignment=Alignment(horizontal="center")
    sm.row_dimensions[r].height=17
sm[f"A{LDIA+3}"].comment=Comment("Fechou ÷ para fechar, dia a dia. O total da "
 "direita é a semana inteira.\n\nDia sem nada previsto fica em branco, não "
 "em zero: zero de nada não é zero por cento.","PCM")
fa_=f"B{LDIA+3}:I{LDIA+3}"
sm.conditional_formatting.add(fa_, CellIsRule(operator="greaterThanOrEqual",
    formula=["0.85"], fill=fill(OK_V), font=F(bold=True,size=10,color=OK_T)))
sm.conditional_formatting.add(fa_, CellIsRule(operator="lessThan",
    formula=["0.6"], fill=fill(RU_V), font=F(bold=True,size=10,color=RU_T)))

LC=LDIA+5
sm[f"A{LC}"]=("Para programar o que está na carteira: na aba Programação, filtre a coluna "
              "Situação por “Na carteira” e escreva a Semana, o Dia e quantos dias leva.")
sm[f"A{LC}"].font=F(size=9,italic=True,color=T2); sm.merge_cells(f"A{LC}:N{LC}")
sm.freeze_panes=f"B{LCAB+1}"; sm.sheet_view.showGridLines=False

# ═══════════════════════════════════ ADERÊNCIA
ad=wb.create_sheet("Aderência")
titulo(ad,"ADERÊNCIA DA PROGRAMAÇÃO  ·  medida por atividade, só da oficina interna","M")
ad["A3"]="Semana:"; ad["A3"].font=F(bold=True,size=10)
ad["B3"]="=Semana!$C$3"; ad["B3"].number_format='"sem "0'
ad["C3"]="=Semana!$E$3"; ad["D3"]="a"; ad["E3"]="=Semana!$G$3"
for ref in ("B3","C3","E3"):
    x=ad[ref]; x.font=F(bold=True,size=11); x.fill=fill(CINZA_C); x.border=box
    x.alignment=Alignment(horizontal="center")
    if ref!="B3": x.number_format=DFMT
ad["D3"].alignment=Alignment(horizontal="center")
ad["F3"]="←  vem da semana escolhida na aba Semana"
ad["F3"].font=F(size=9,italic=True,color=T2); ad.merge_cells("F3:M3")
janA=f'{INI},">="&$C$3,{INI},"<="&$E$3'
janW=f'{PLANOD},">="&$C$3,{PLANOD},"<="&$E$3'

cabec(ad,11,[("A","Indicador"),("B","Valor"),("C","Como é medido")],NAVY)
for _c in "DEFGHIJKLM":
    x=ad[f"{_c}11"]; x.fill=fill(NAVY); x.border=box
ad.merge_cells("C11:M11")
ad["C11"].alignment=Alignment(horizontal="left",vertical="center",indent=1)

SEC="__sec__"
IND=[
 (SEC,"O QUE A OFICINA INTERNA TINHA PARA FAZER","",""),
 ("Atividades do plano", f'=COUNTIFS({janA},{PLAN})',"0",
  "estavam programadas para a semana"),
 ("Extra programação", f'=COUNTIFS({janA},{ORIG},"Extra",{BASE})',"0",
  "entraram depois: quebra, urgência, pedido da operação"),
 ("Concluídas do plano", f'=SUMIFS({CONCL},{janA},{NEX},{INT})',"0",
  "do que estava programado, quanto saiu"),
 ("Total interno", f'=COUNTIFS({janA},{BASE})',"0","plano + extra"),
 ("Extras concluídas", f'=SUMIFS({CONCL},{janA},{ORIG},"Extra",{INT})',"0",""),
 ("Total concluído", f'=SUMIFS({CONCL},{janA},{INT})',"0","plano + extra"),
 (SEC,"OS NÚMEROS","",""),
 ("Aderência à programação", '=IFERROR({Concluídas do plano}/{Atividades do plano},"")',"0%",
  "concluídas do plano ÷ atividades do plano — o indicador principal"),
 ("Cumprimento geral", '=IFERROR({Total concluído}/{Total interno},"")',"0%",
  "total concluído ÷ total interno — inclui o extra"),
 ("Quanto da semana foi extra", '=IFERROR({Extra programação}/{Total interno},"")',"0%",
  "extra ÷ total interno"),
 ("Concluídas no prazo", f'=IF(COUNTIFS({janA},{CONCLEM},"<>")=0,"",SUMIFS({NOPRAZO},{janA},{NEX},{INT}))',"0",
  "saíram no próprio dia programado"),
 ("Pontualidade", f'=IFERROR({{Concluídas no prazo}}/COUNTIFS({janA},{CONCLEM},"<>",{PLAN}),"")',"0%",
  "no prazo ÷ concluídas com data"),
 ("Aderência ao plano original", f'=IF(COUNTIFS({janW},{CONCLEM},"<>")=0,"",'
  f'IFERROR(SUMIFS({NOPLANO},{janW},{NEX},{INT})/COUNTIFS({janW},{PLAN}),""))',"0%",
  "contra a semana da 1ª programação — não melhora quando se empurra para a frente"),
 ("Saíram desta semana", f'=COUNTIFS({SEMORIG},$B$3,{SEM},"<>"&$B$3,{NC})',"0",
  "estavam programadas para esta semana e foram empurradas para outra. A "
  "aderência acima não cobra mais elas — quem cobra é a aderência ao plano "
  "original. Quando este número é grande, é ela que conta a verdade da semana"),
 ("Vencidas", f'=COUNTIFS({SIT},"VENCIDA",{janA},{INT})',"0",
  "a janela de execução fechou e a atividade não saiu"),
 ("Atividades reprogramadas", f'=SUMIFS({REP},{janA},{INT})',"0",
  "mudaram de semana"),
 ("Atraso médio, quando atrasa", f'=IFERROR(SUMIFS({ATRASO},{janA},{INT})/COUNTIFS({janA},{ATRASO},">0",{INT}),0)',"0.0",
  "dias entre o dia programado e a conclusão"),
 (SEC,"DIAS DE SERVIÇO","",""),
 ("Diária — atividades por dia útil", '=IFERROR({Total interno}/5,"")',"0.0",
  "total interno ÷ 5 dias — quantas atividades a semana pede por dia"),
 ("Diária do que saiu", '=IFERROR({Total concluído}/5,"")',"0.0",
  "total concluído ÷ 5 dias — o que a oficina de fato entrega por dia"),
 ("Soma das estimativas (dias)", f'=SUMIFS({DIAS},{janA},{BASE})',"0",
  "soma da coluna Dias prev. Como a estimativa é a duração do SERVIÇO e cada "
  "atividade dele repete o número, este total infla quando o serviço tem muitas "
  "atividades — use-o para comparar semanas, não como homem-dia"),
 (SEC,"FORA DA CONTA INTERNA","",""),
 ("Em empresa terceirizada", f'=COUNTIFS({janA},{OFIC},"Terceirizada",{NC})',"0",
  "serviço de fora: não mede a oficina da Makro"),
 ("Canceladas", f'=COUNTIFS({janA},{MARC},"Cancelada")',"0","saem dos dois lados"),
 ("Na carteira, sem dia", f'=COUNTIFS({INI},"",{ATIV},"<>",{NC})',"0",
  "não depende da semana — é o que ainda espera encaixe"),
]
# Cada indicador cita os outros pelo NOME, não pela linha: acrescentar uma
# linha no meio da tabela deixava de quebrar as contas de baixo.
LINHA_DE={}
_r=12
for _rot,_f,_fm,_e in IND:
    if _rot!=SEC: LINHA_DE[_rot]=f"B{_r}"
    _r+=1
def resolver(f):
    if not isinstance(f,str) or "{" not in f: return f
    for nome,end in LINHA_DE.items(): f=f.replace("{"+nome+"}",end)
    return f
HERO=[("A","F","ADERÊNCIA À PROGRAMAÇÃO",'=IFERROR({Concluídas do plano}/{Atividades do plano},"")',
       '="do plano da semana, "&{Concluídas do plano}&" de "&{Atividades do plano}&" atividades foram concluídas."&CHAR(10)&'
       '"Extra programação, serviço de terceiro e cancelada ficam fora dos dois lados."',NAVY),
      ("G","M","CUMPRIMENTO GERAL",'=IFERROR({Total concluído}/{Total interno},"")',
       '="contando o extra, a oficina fez "&{Total concluído}&" das "&{Total interno}&" atividades que teve."&CHAR(10)&'
       '"A distância entre os dois números é o tamanho do imprevisto na semana."',AZ_T)]
for c1,c2,rot,fml,txt,cor in HERO:
    ad.merge_cells(f"{c1}5:{c2}5"); ad.merge_cells(f"{c1}6:{c2}7"); ad.merge_cells(f"{c1}8:{c2}9")
    a=ad[f"{c1}5"]; a.value=rot
    a.font=F(bold=True,size=9.5,color=BRANCO); a.fill=fill(NAVY)
    a.alignment=Alignment(horizontal="center",vertical="center")
    b=ad[f"{c1}6"]
    b.font=F(bold=True,size=34,color=cor); b.number_format="0%"
    b.alignment=Alignment(horizontal="center",vertical="center"); b.fill=fill(CINZA_C)
    d=ad[f"{c1}8"]
    d.font=F(size=10,color=T2)
    d.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    d.fill=fill(CINZA_C)
    for rr in range(5,10):
        for cc in [get_column_letter(x) for x in range(ord(c1)-64, ord(c2)-64+1)]:
            ad[f"{cc}{rr}"].border=box
for rr,h in ((5,18),(6,26),(7,26),(8,16),(9,18)): ad.row_dimensions[rr].height=h

HERO=[(a,b,rot,resolver(fml),resolver(txt),cor) for a,b,rot,fml,txt,cor in HERO]
for c1,c2,rot,fml,txt,cor in HERO:
    ad[f"{c1}6"]=fml; ad[f"{c1}8"]=txt

r=12
for rot,fml,fmt,expl in IND:
    fml=resolver(fml)
    if rot==SEC:
        ad[f"A{r}"]=fml
        ad[f"A{r}"].font=F(bold=True,size=9,color=BRANCO); ad[f"A{r}"].fill=fill(NAVY2)
        ad[f"A{r}"].alignment=Alignment(vertical="center",indent=1)
        ad.merge_cells(f"A{r}:M{r}")
        for cc in "ABCDEFGHIJKLM": ad[f"{cc}{r}"].border=box
        ad.row_dimensions[r].height=17; r+=1; continue
    ad[f"A{r}"]=rot; ad[f"B{r}"]=fml; ad[f"C{r}"]=expl
    dest=rot in ("Aderência à programação","Cumprimento geral")
    ad[f"A{r}"].font=F(size=10,bold=dest)
    ad[f"B{r}"].font=F(size=11,bold=True,color=NAVY if dest else TINTA)
    ad[f"B{r}"].number_format=fmt
    ad[f"B{r}"].alignment=Alignment(horizontal="center")
    ad[f"C{r}"].font=F(size=9,color=T2)
    for cc in "ABC": ad[f"{cc}{r}"].border=box
    if dest:
        for cc in "ABC": ad[f"{cc}{r}"].fill=fill("FFEDF1F9")
    ad.merge_cells(f"C{r}:M{r}")
    ad.row_dimensions[r].height=17
    r+=1
FIM_IND=r
for rot in ("Aderência à programação","Cumprimento geral","Pontualidade",
            "Aderência ao plano original"):
    ref=LINHA_DE[rot]
    ad.conditional_formatting.add(ref, CellIsRule(operator="greaterThanOrEqual",
        formula=["0.85"], font=F(bold=True,size=11,color=OK_T)))
    ad.conditional_formatting.add(ref, CellIsRule(operator="lessThan",
        formula=["0.6"], font=F(bold=True,size=11,color=RU_T)))
ad.conditional_formatting.add(LINHA_DE["Vencidas"], CellIsRule(operator="greaterThan",
    formula=["0"], fill=fill(RU_V), font=F(bold=True,size=11,color=RU_T)))
for rot in ("Extra programação","Quanto da semana foi extra"):
    ad.conditional_formatting.add(LINHA_DE[rot], CellIsRule(operator="greaterThan",
        formula=["0"], fill=fill(EX_V), font=F(bold=True,size=11,color=EX_T)))
for a6,rot in (("A6","Aderência à programação"),("G6","Cumprimento geral")):
    ad.conditional_formatting.add(a6, CellIsRule(operator="greaterThanOrEqual",
        formula=["0.85"], font=F(bold=True,size=34,color=OK_T)))
    ad.conditional_formatting.add(a6, CellIsRule(operator="lessThan",
        formula=["0.6"], font=F(bold=True,size=34,color=RU_T)))

# ═══════════════════════════════════ GRÁFICOS
# Dois gráficos, um ao lado do outro: à esquerda o tamanho de cada bloco e
# quanto dele saiu; à direita onde a semana parou. Os números que os
# alimentam ficam nas colunas O:Q, fora da área de impressão — mexer neles
# é mexer nos indicadores de cima, que é de onde eles vêm.
def _txt(tam=9, negrito=False, cor=T2):
    """Fonte dos rótulos do gráfico — o padrão do Excel sai pequeno demais."""
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(
        sz=int(tam*100), b=negrito, latin=None,
        solidFill=cor[2:])), endParaRPr=CharacterProperties(sz=int(tam*100)))],
        bodyPr=RichTextProperties())

DAD=FIM_IND+1                       # canto do bloco de números dos gráficos
ad[f"O{DAD}"]="Bloco"; ad[f"P{DAD}"]="Tinha para fazer"; ad[f"Q{DAD}"]="Concluídas"
ad[f"O{DAD+1}"]="Plano da semana"
ad[f"P{DAD+1}"]=f"={LINHA_DE['Atividades do plano']}"
ad[f"Q{DAD+1}"]=f"={LINHA_DE['Concluídas do plano']}"
ad[f"O{DAD+2}"]="Extra programação"
ad[f"P{DAD+2}"]=f"={LINHA_DE['Extra programação']}"
ad[f"Q{DAD+2}"]=f"={LINHA_DE['Extras concluídas']}"

E0=DAD+4
ad[f"O{E0}"]="Situação"; ad[f"P{E0}"]="Atividades"
SITG=[("Concluídas",  f"={LINHA_DE['Total concluído']}",                    OK_T),
      ("Vencidas",    f"={LINHA_DE['Vencidas']}",                           RU_T),
      ("A vencer",    f"=MAX(0,{LINHA_DE['Total interno']}-{LINHA_DE['Total concluído']}"
                      f"-{LINHA_DE['Vencidas']})",                          "FF2C4272"),
      ("Canceladas",  f"={LINHA_DE['Canceladas']}",                         "FF98A1B2"),
      ("Terceirizada",f"={LINHA_DE['Em empresa terceirizada']}",            EXT)]
for k,(rot,fml,_c) in enumerate(SITG):
    ad[f"O{E0+1+k}"]=rot; ad[f"P{E0+1+k}"]=fml
for rr in list(range(DAD,DAD+3))+list(range(E0,E0+1+len(SITG))):
    for cc in "OPQ":
        ad[f"{cc}{rr}"].font=F(size=9,color=T2)
for cc,w in (("O",18),("P",15),("Q",12)): ad.column_dimensions[cc].width=w

CH=FIM_IND+1                        # onde os desenhos ficam ancorados

g1=BarChart(); g1.type="col"; g1.grouping="clustered"; g1.gapWidth=60; g1.overlap=-15
g1.title="Quanto tinha, quanto saiu"
g1.add_data(Reference(ad,min_col=16,max_col=17,min_row=DAD,max_row=DAD+2),
            titles_from_data=True)
g1.set_categories(Reference(ad,min_col=15,min_row=DAD+1,max_row=DAD+2))
g1.series[0].graphicalProperties=GraphicalProperties(solidFill=NAVY[2:])
g1.series[1].graphicalProperties=GraphicalProperties(solidFill=OK_T[2:])

g2=BarChart(); g2.type="col"; g2.grouping="clustered"; g2.gapWidth=45
g2.title="Onde a semana parou"
g2.add_data(Reference(ad,min_col=16,min_row=E0,max_row=E0+len(SITG)),
            titles_from_data=True)
g2.set_categories(Reference(ad,min_col=15,min_row=E0+1,max_row=E0+len(SITG)))
g2.legend=None
# cada barra na cor que ela já tem na planilha: verde concluída, vermelho
# vencida, azul a vencer, cinza cancelada, laranja terceirizada.
g2.series[0].data_points=[DataPoint(idx=k,
    spPr=GraphicalProperties(solidFill=cor[2:]))
    for k,(_r,_f,cor) in enumerate(SITG)]

for g,anc,larg in ((g1,f"A{CH}",14.5),(g2,f"H{CH}",11.5)):
    g.height=8.2; g.width=larg
    g.y_axis.majorGridlines=None
    g.y_axis.title=None; g.x_axis.title=None
    g.y_axis.numFmt="0"
    g.dLbls=DataLabelList(); g.dLbls.showVal=True; g.dLbls.showLegendKey=False
    g.dLbls.showCatName=False; g.dLbls.showSerName=False
    g.dLbls.txPr=_txt(9,True,TINTA)
    g.x_axis.txPr=_txt(9); g.y_axis.txPr=_txt(8)
    g.txPr=_txt(9)
    if g.legend is not None:
        g.legend.position="b"; g.legend.overlay=False; g.legend.txPr=_txt(9)
    ad.add_chart(g,anc)
CH_FIM=CH+16

# ── recortes ──
def recorte(lin, col0, tit, lista, coluna, n=22, equipe=False):
    cols=[get_column_letter(col0+i) for i in range(6)]
    a,b,c,d,e,g=cols
    ad[f"{a}{lin}"]=tit
    ad[f"{a}{lin}"].font=F(bold=True,size=11,color=NAVY)
    ad.merge_cells(f"{a}{lin}:{g}{lin}")
    cabec(ad,lin+1,[(a,"Nome"),(b,"Plano"),(c,"Concl."),(d,"Aderência"),
                    (e,"Vencidas"),(g,"Extra")])
    for k in range(n):
        r=lin+2+k
        ad[f"{a}{r}"]=f'=IF(Listas!{lista}{5+k}="","",Listas!{lista}{5+k})'
        venc=SIT+',"VENCIDA",'+janA+","+INT
        extra=f'{janA},{ORIG},"Extra",{BASE}'
        if equipe:
            fb=porpessoa(f"${a}{r}", f'{janA},{PLAN}')
            fc=porpessoa(f"${a}{r}", f'{janA},{NEX},{INT}', "SUMIFS", CONCL)
            fe=porpessoa(f"${a}{r}", venc)
            fg=porpessoa(f"${a}{r}", extra)
        else:
            fb=f'COUNTIFS({coluna},${a}{r},{janA},{PLAN})'
            fc=f'SUMIFS({CONCL},{coluna},${a}{r},{janA},{NEX},{INT})'
            fe=f'COUNTIFS({coluna},${a}{r},{venc})'
            fg=f'COUNTIFS({coluna},${a}{r},{extra})'
        for cl,fm in ((b,fb),(c,fc),(e,fe),(g,fg)):
            ad[f"{cl}{r}"]=f'=IF(${a}{r}="","",{fm})'
        ad[f"{d}{r}"]=f'=IF(OR(${a}{r}="",${b}{r}=0),"",${c}{r}/${b}{r})'
        for col in cols:
            cc=ad[f"{col}{r}"]; cc.border=box; cc.font=F(size=10)
            if col!=a: cc.alignment=Alignment(horizontal="center")
            if col==d: cc.number_format="0%"
            if col in (b,c,d,e,g): cc.fill=fill(CINZA_C)
        ad.row_dimensions[r].height=16
    ad.conditional_formatting.add(f"{g}{lin+2}:{g}{lin+1+n}", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=fill(EX_V), font=F(bold=True,size=10,color=EX_T)))
    fa=f"{d}{lin+2}:{d}{lin+1+n}"
    ad.conditional_formatting.add(fa, CellIsRule(operator="greaterThanOrEqual",
        formula=["0.85"], fill=fill(OK_V), font=F(bold=True,size=10,color=OK_T)))
    ad.conditional_formatting.add(fa, CellIsRule(operator="lessThan",
        formula=["0.6"], fill=fill(RU_V), font=F(bold=True,size=10,color=RU_T)))
    ad.conditional_formatting.add(f"{e}{lin+2}:{e}{lin+1+n}", CellIsRule(
        operator="greaterThan", formula=["0"], font=F(bold=True,size=10,color=RU_T)))
    if not equipe: return
    # A coluna acima soma mais que a semana inteira, e é isso mesmo: atividade
    # dividida conta cheia para cada um que participou. O rodapé traz o número
    # de atividades DISTINTAS, para os dois se lerem lado a lado sem parecer
    # erro de conta. A razão de cada pessoa continua certa — é dela que sai a
    # aderência individual.
    rt=lin+2+n
    ad[f"{a}{rt}"]="TOTAL — atividades distintas"
    ad[f"{b}{rt}"]=f'=COUNTIFS({janA},{PLAN})'
    ad[f"{c}{rt}"]=f'=SUMIFS({CONCL},{janA},{NEX},{INT})'
    ad[f"{d}{rt}"]=f'=IF(${b}{rt}=0,"",${c}{rt}/${b}{rt})'
    ad[f"{e}{rt}"]=f'=COUNTIFS({SIT},"VENCIDA",{janA},{INT})'
    ad[f"{g}{rt}"]=f'=COUNTIFS({janA},{ORIG},"Extra",{BASE})'
    for col in cols:
        cc=ad[f"{col}{rt}"]; cc.border=box
        cc.font=F(size=10,bold=True,color=BRANCO); cc.fill=fill(NAVY2)
        if col!=a: cc.alignment=Alignment(horizontal="center")
        if col==d: cc.number_format="0%"
    ad[f"{a}{rt}"].comment=Comment("Cada atividade conta CHEIA para cada um que "
      "participou, então a coluna acima soma mais que o total da semana.\n\n"
      "Aqui está o número de atividades distintas. O peso rateado — a atividade "
      "dividida entre quem participou — fica na coluna PESO da aba Semana.","PCM")
    ad.row_dimensions[rt].height=17

L0=CH_FIM+2
N_EX=max(14,len(EXEC_LISTA)+3); N_FR=max(14,len(FROTA_LISTA)+3)
recorte(L0, 1, "POR EXECUTANTE  ·  atividade dividida conta para cada um",
        "A", None, N_EX, equipe=True)
recorte(L0, 8, "POR FROTA", "B", FROTA_, N_FR)
recorte(L0+max(N_EX,N_FR)+4, 1, "POR TIPO DE SERVIÇO", "C", TIPO_, len(TIPO_LISTA)+2)
ad.column_dimensions["A"].width=30
for col,w in (("B",9),("C",9),("D",11),("E",10),("F",9),("G",3),
              ("H",15),("I",9),("J",9),("K",11),("L",10),("M",9)):
    ad.column_dimensions[col].width=w
ad.sheet_view.showGridLines=False; ad.freeze_panes="A4"

# ═══════════════════════════════════ HOJE
# O painel do dia, e a folha que se imprime e leva para a oficina. Ele não
# guarda nada: tudo vem da Programação pelas colunas de ordem, e a data lá em
# cima é o único campo que se digita.
hj=wb.create_sheet("Hoje")
titulo(hj,"O DIA  ·  o que a oficina tem para fazer e o que já fechou","I")
hj["A3"]="DATA"; hj["A3"].font=F(bold=True,size=11,color=NAVY)
hj["A3"].alignment=Alignment(horizontal="right",vertical="center")
hj.merge_cells("A3:B3")
hj["C3"]="=TODAY()"
x=hj["C3"]; x.font=F(bold=True,size=14,color=NAVY); x.fill=fill(AMARELO)
x.border=box; x.alignment=Alignment(horizontal="center",vertical="center")
x.number_format=DFMT
x.comment=Comment("Vem preenchida com o dia de hoje.\n\nTroque a data para "
 "olhar outro dia — a lista, os indicadores e as atrasadas acompanham. "
 "Para voltar ao normal, escreva =HOJE()","PCM")
hj["D3"]='=IF($C$3="","",TEXT($C$3,"[$-416]dddd"))'
hj["D3"].font=F(bold=True,size=11,color=T2)
hj["D3"].alignment=Alignment(horizontal="center",vertical="center")
# ISOWEEKNUM não existe em toda build. O número sai da mesma âncora que a
# coluna Início usa — assim a semana daqui e a de lá nunca divergem.
hj["E3"]=f'=IF($C$3="","","semana "&(INT(($C$3-{SEG})/7)+1))'
hj["E3"].font=F(bold=True,size=11,color=T2)
hj["E3"].alignment=Alignment(horizontal="center",vertical="center")
hj["F3"]="←  troque a data para ver outro dia"
hj["F3"].font=F(size=9,italic=True,color=T2); hj.merge_cells("F3:I3")
hj["F3"].alignment=Alignment(vertical="center")
hj.row_dimensions[3].height=26

# ── bloco motor ──
# Fica em K:L, fora da área de impressão. As duas primeiras células são as
# únicas voláteis da planilha inteira: antes TODAY() e NOW() moravam dentro
# das mil linhas da Programação e o Excel refazia cinco mil chamadas a cada
# tecla digitada.
hj["K1"]="motor — não apague"
hj["K1"].font=F(bold=True,size=8,color=RU_T)
for lin,rot,fml,fmt in ((1,"hoje","=TODAY()",DFMT),
                        (2,"agora","=NOW()-TODAY()","hh:mm"),
                        (3,"no dia",f'=SUM({rg("AH")})',"0"),
                        (4,"atrasadas",f'=SUM({rg("AI")})',"0")):
    if lin>1:
        hj[f"K{lin}"]=rot; hj[f"K{lin}"].font=F(size=8,color=T2)
    hj[f"L{lin}"]=fml
    hj[f"L{lin}"].font=F(size=8,color=T2); hj[f"L{lin}"].number_format=fmt
hj["M1"]=("Estas quatro células movem a planilha: a data de hoje, a hora, e o "
          "tamanho das duas listas abaixo. Não apague nem mova.")
hj["M1"].font=F(size=8,italic=True,color=T2)
for c,w in (("K",13),("L",10),("M",60)): hj.column_dimensions[c].width=w

D_="$C$3"
janD=f'{FIM},{D_}'
RESH=[("A","B","PARA FECHAR HOJE",f'=COUNTIFS({janD},{BASE})',"0",NAVY,
       "atividades cujo término previsto cai neste dia"),
      ("C","C","FECHARAM",f'=SUMIFS({CONCL},{janD},{INT},{NC})',"0",OK_T,
       "dessas, quantas já estão concluídas"),
      ("D","E","ADERÊNCIA DO DIA",'=IF($A$6=0,"",$C$6/$A$6)',"0%",NAVY,
       "fecharam ÷ para fechar — é a aderência diária, por atividade"),
      ("F","G","EM EXECUÇÃO",
       f'=COUNTIFS({INI},"<="&{D_},{FIM},">="&{D_},{CONCL},0,{BASE})',"0",AZ_T,
       "abertas e dentro da janela de execução neste dia"),
      ("H","I","ATRASADAS",f'=COUNTIFS({FIM},"<"&{D_},{CONCL},0,{BASE})',"0",RU_T,
       "a janela fechou e a atividade não saiu — vêm de dias anteriores")]
for c1,c2,rot,fml,fmt,cor,dica in RESH:
    hj.merge_cells(f"{c1}5:{c2}5"); hj.merge_cells(f"{c1}6:{c2}6")
    a=hj[f"{c1}5"]; a.value=rot
    a.font=F(bold=True,size=8.5,color=T2); a.fill=fill(CINZA_C)
    a.alignment=Alignment(horizontal="center",wrap_text=True)
    a.comment=Comment(dica,"PCM")
    b=hj[f"{c1}6"]; b.value=fml
    b.font=F(bold=True,size=20,color=cor); b.number_format=fmt
    b.alignment=Alignment(horizontal="center",vertical="center")
    b.fill=fill(RU_V if rot=="ATRASADAS" else CINZA_C)
    for cc in {c1,c2}:
        hj[f"{cc}5"].border=box; hj[f"{cc}6"].border=box
hj.row_dimensions[5].height=22; hj.row_dimensions[6].height=30
hj.conditional_formatting.add("D6", CellIsRule(operator="greaterThanOrEqual",
    formula=["0.85"], font=F(bold=True,size=20,color=OK_T)))
hj.conditional_formatting.add("D6", CellIsRule(operator="lessThan",
    formula=["0.6"], font=F(bold=True,size=20,color=RU_T)))
hj.conditional_formatting.add("H6", CellIsRule(operator="equal",
    formula=["0"], fill=fill(OK_V), font=F(bold=True,size=20,color=OK_T)))

COLS_H=[("A","Frota",10),("B","Serviço",28),("C","Atividade",40),
        ("D","Quem faz",26),("E","Início",11),("F","Fim",11),
        ("G","Situação",17),("H","Concluída em",12),("I","Linha",7)]
for col,_t,w in COLS_H: hj.column_dimensions[col].width=w
DE_ONDE=[("A","D"),("B","E"),("C","F"),("D","AG"),("E","O"),("F","P"),("H","Q")]

def bloco(lin, rot, sub, ordem, total, n, cor):
    """Uma lista puxada da Programação pela coluna de ordem, sem matricial."""
    hj[f"A{lin}"]=rot
    hj[f"A{lin}"].font=F(bold=True,size=11,color=cor)
    hj[f"A{lin}"].alignment=Alignment(vertical="center",indent=1)
    hj.merge_cells(f"A{lin}:C{lin}")
    hj[f"D{lin}"]=sub
    hj[f"D{lin}"].font=F(size=9,italic=True,color=T2)
    hj.merge_cells(f"D{lin}:I{lin}")
    hj[f"D{lin}"].alignment=Alignment(vertical="center")
    hj.row_dimensions[lin].height=20
    cabec(hj,lin+1,[(c,t) for c,t,_w in COLS_H])
    # Quantas linhas a lista tem, vindo do bloco motor. Antes era a última
    # célula da coluna de ordem, que só servia enquanto a ordem era um
    # contador em cadeia — e a cadeia é justamente o que causou os #REF!.
    fim_=total
    for k in range(1,n+1):
        r=lin+1+k
        pos=f'MATCH({k},{ordem},0)'
        for col,orig in DE_ONDE:
            alvo=f"INDEX({rg(orig)},{pos})"
            # Célula vazia puxada por INDEX vale zero, e zero com formato de
            # data vira 30/12/1899 na tela. A data de conclusão é a única que
            # pode estar vazia numa linha listada, então ela leva a guarda.
            if col=="H": alvo=f'IF({alvo}="","",{alvo})'
            hj[f"{col}{r}"]=f'=IF({fim_}<{k},"",{alvo})'
        hj[f"I{r}"]=f'=IF({fim_}<{k},"",{pos}+{PRIM-1})'
        # situação relativa à DATA escolhida, não a hoje: trocar a data no
        # topo tem de mudar a leitura da lista inteira.
        hj[f"G{r}"]=(f'=IF($A{r}="","",IF($H{r}<>"","Concluída",'
                     f'IF($F{r}<$C$3,"Atrasada",'
                     f'IF($F{r}=$C$3,"Fecha neste dia","Em execução"))))')
        for col,_t,_w in COLS_H:
            c=hj[f"{col}{r}"]; c.border=box; c.font=F(size=10)
            if col in ("E","F","H"): c.number_format=DFMT
            if col in ("A","E","F","G","H","I"): c.alignment=Alignment(horizontal="center")
            if col=="C": c.alignment=Alignment(vertical="center",indent=1)
        hj.row_dimensions[r].height=16
    # A lista tem teto. Sem este aviso um dia cheio truncaria calado, e ele
    # tocaria a oficina achando que viu tudo.
    hj[f"A{lin+2+n}"]=(f'=IF({fim_}<={n},"",'
      f'"⚠ há mais "&({fim_}-{n})&" atividades além das {n} listadas — '
      f'veja na aba Programação")')
    hj[f"A{lin+2+n}"].font=F(bold=True,size=10,color=RU_T)
    hj[f"A{lin+2+n}"].alignment=Alignment(vertical="center",indent=1)
    hj.merge_cells(f"A{lin+2+n}:I{lin+2+n}")
    hj.row_dimensions[lin+2+n].height=18
    faixa=f"G{lin+2}:G{lin+1+n}"
    for txt,bg,fg in (("Concluída",OK_V,OK_T),("Fecha neste dia",EX_V,EX_T),
                      ("Atrasada",RU_V,RU_T),("Em execução",AZ_V,AZ_T)):
        rr=Rule(type="expression", stopIfTrue=True,
                dxf=DifferentialStyle(fill=PatternFill(bgColor=bg),
                                      font=Font(color=fg,bold=True)))
        rr.formula=[f'$G{lin+2}="{txt}"']
        hj.conditional_formatting.add(faixa, rr)
    return lin+3+n

N_DIA=45; N_ATR=40
# "Para fechar hoje" é zero em todo dia que não é término de nada — numa
# segunda com estimativas de três dias, o painel inteiro ficava mudo. O que
# de fato saiu no dia entra aqui, e essa linha responde todo dia.
L1=bloco(8,"O DIA",
         f'="a janela de execução contém esta data   ·   CONCLUÍDAS NESTA '
         f'DATA: "&COUNTIFS({CONCLEM},{D_},{BASE})',
         ORD_DIA,"$L$3",N_DIA,NAVY)
L2=bloco(L1+1,"ATRASADAS","a janela já fechou e a atividade continua aberta — "
         "estas vêm de dias anteriores",ORD_ATR,"$L$4",N_ATR,RU_T)
hj[f"A{L2+1}"]=("Para concluir uma atividade: vá à aba PROGRAMAÇÃO, na linha "
  "que a coluna LINHA indica, e escreva a data em CONCLUÍDA EM. Aqui nada se "
  "digita a não ser a data lá em cima — o resto é espelho da Programação.")
hj[f"A{L2+1}"].font=F(size=9,italic=True,color=T2)
hj[f"A{L2+1}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
hj.merge_cells(f"A{L2+1}:I{L2+1}"); hj.row_dimensions[L2+1].height=28
hj.sheet_view.showGridLines=False; hj.freeze_panes="A10"

# ═══════════════════════════════════ COMO USAR
ins=wb.create_sheet("Como usar", 0)
titulo(ins,"PROGRAMAÇÃO DE SERVIÇOS  ·  MAKRO TRANSPORTES","H",34,15)
ins["A2"]="Planejamento e Controle de Manutenção  ·  aderência diária, por atividade"
ins["A2"].font=F(size=10,color=BRANCO); ins["A2"].fill=fill(NAVY2)
ins["A2"].alignment=Alignment(vertical="center",indent=1)
ins.merge_cells("A2:H2"); ins.row_dimensions[2].height=20
def sec(r,txt):
    ins[f"A{r}"]=txt
    ins[f"A{r}"].font=F(bold=True,size=11,color=NAVY)
    ins.merge_cells(f"A{r}:H{r}"); ins.row_dimensions[r].height=24
def par(r,txt,alt=None,cor=TINTA):
    ins[f"A{r}"]=txt
    ins[f"A{r}"].font=F(size=10,color=cor)
    ins[f"A{r}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
    ins.merge_cells(f"A{r}:H{r}")
    if alt: ins.row_dimensions[r].height=alt
def passo(r,n,txt,cor=NAVY):
    ins[f"A{r}"]=n
    ins[f"A{r}"].font=F(bold=True,size=13,color=BRANCO); ins[f"A{r}"].fill=fill(cor)
    ins[f"A{r}"].alignment=Alignment(horizontal="center",vertical="center")
    ins[f"A{r}"].border=box
    ins[f"B{r}"]=txt; ins[f"B{r}"].font=F(size=10.5)
    ins[f"B{r}"].alignment=Alignment(vertical="center",wrap_text=True,indent=1)
    ins.merge_cells(f"B{r}:H{r}")
    for col in "BCDEFGH": ins[f"{col}{r}"].border=box
    ins.row_dimensions[r].height=34

L=4
sec(L,"O DIA A DIA — comece pela aba HOJE"); L+=1
par(L,"A aba HOJE abre no dia de hoje e responde três coisas: o que tem de "
      "FECHAR hoje, o que já fechou, e o que ficou atrasado de dias "
      "anteriores. A ADERÊNCIA DO DIA é fechou ÷ para fechar — a mesma conta "
      "da semana, medida no dia.",alt=42); L+=1
par(L,"A lista embaixo é o que está em execução: a janela de cada atividade "
      "vai do DIA dela até o TÉRMINO PREVISTO. Enquanto a janela está aberta a "
      "situação é EM EXECUÇÃO; no último dia vira FECHA HOJE; só depois disso "
      "ela fica VENCIDA.",alt=42); L+=1
par(L,"Para CONCLUIR: a coluna LINHA da aba Hoje diz a linha exata na "
      "Programação. Vá lá e escreva a data em CONCLUÍDA EM. Na aba Hoje só se "
      "digita a data lá em cima — o resto é espelho.",alt=36,cor=T2); L+=2

sec(L,"MONTAR A SEMANA — quatro passos"); L+=1
passo(L,"1","Na aba PROGRAMAÇÃO, filtre a coluna SITUAÇÃO por “Na carteira”. "
           "Sobram as atividades que ainda não têm dia."); L+=1
passo(L,"2","Escreva o NÚMERO DA SEMANA (35, 36…) na coluna Semana. Numa linha "
           "e arraste para baixo — vale para todas."); L+=1
passo(L,"3","Escolha o DIA, quantos DIAS PREV. o serviço leva, e quem faz em "
           "EXECUTANTE 1, 2 e 3. Início, término previsto e Situação saem sozinhos."); L+=1
passo(L,"4","Volte à aba SEMANA: ela mostra a carga de cada um por dia e os "
           "dias de serviço — a diária."); L+=2

sec(L,"MEXER NAS LINHAS — no Excel"); L+=1
par(L,"ACRESCENTAR NO FIM: clique na última linha preenchida, pegue a alça no "
      "canto da seleção e puxe para baixo. As fórmulas vêm junto.",alt=30); L+=1
par(L,"INSERIR NO MEIO: selecione uma LINHA INTEIRA parecida, Ctrl+C, botão "
      "direito na linha de baixo e “Inserir células copiadas”. Não use "
      "“Inserir linha” pura — ela entra sem fórmula nenhuma e fica muda: sem "
      "Nº, sem Situação, sem data.",alt=42); L+=1
par(L,"APAGAR: apague a LINHA INTEIRA, pelo número dela à esquerda. Pode "
      "apagar à vontade — as contas se ajustam sozinhas.",alt=30); L+=2

sec(L,"TROCAR DE SEMANA"); L+=1
par(L,"Na aba SEMANA, a célula amarela grande é o NÚMERO da semana. Troque e tudo "
      "acompanha. Ao lado ficam o ANO e a hora em que a oficina fecha.",alt=28); L+=1
par(L,"A faixa ONDE TEM PROGRAMAÇÃO mostra quantas atividades existem nas quatro "
      "semanas antes e nas quatro depois.",alt=26); L+=2

sec(L,"QUANTOS DIAS O SERVIÇO LEVA"); L+=1
par(L,"A coluna DIAS PREV. é a sua estimativa: 1 começa e acaba no mesmo dia, 3 leva "
      "três dias. Dela sai o TÉRMINO PREVISTO — uma atividade de três dias começada "
      "na segunda só fica vencida depois da quarta.",alt=32); L+=1
par(L,"A soma dessas estimativas é a DIÁRIA: quantos dias de oficina a semana pede. "
      "Aparece por pessoa na aba Semana e no total na aba Aderência.",alt=28); L+=1
par(L,"As atividades já vieram com uma estimativa de dois a três dias, pelo tamanho do "
      "serviço. É um ponto de partida — ajuste linha a linha conforme a realidade.",
      alt=28,cor=T2); L+=2

sec(L,"O QUE ENTRA E O QUE NÃO ENTRA NA ADERÊNCIA"); L+=1
ins[f"A{L}"]="Aderência à programação  =  concluídas do plano  ÷  atividades do plano"
ins[f"A{L}"].font=F(bold=True,size=12,color=NAVY); ins[f"A{L}"].fill=fill("FFEDF1F9")
ins[f"A{L}"].alignment=Alignment(vertical="center",indent=1)
ins.merge_cells(f"A{L}:H{L}"); ins.row_dimensions[L].height=26
for col in "ABCDEFGH": ins[f"{col}{L}"].border=box
L+=1
for tit,txt in [
 ("Empresa terceirizada","Fica FORA. É serviço de fora, não mede a oficina da Makro. "
  "Marque escrevendo “(externo)” no fim do nome do executante."),
 ("Extra programação","Fica FORA da aderência: não dá para cobrar cumprimento de uma "
  "coisa que nunca foi programada. Tem contador próprio."),
 ("Cancelada","Sai dos dois lados da divisão."),
 ("Cumprimento geral","Ao lado da aderência fica este segundo número, que INCLUI o "
  "extra. A distância entre os dois é o tamanho do imprevisto na semana."),
 ("Vencida","Passou do término previsto sem conclusão. Depois da hora de fechar a "
  "oficina, o que vencia hoje já entra como vencida — não espera o dia seguinte.")]:
    ins[f"A{L}"]=tit; ins[f"A{L}"].font=F(bold=True,size=10,color=NAVY)
    ins[f"A{L}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
    ins.merge_cells(f"A{L}:B{L}")
    ins[f"C{L}"]=txt; ins[f"C{L}"].font=F(size=10)
    ins[f"C{L}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
    ins.merge_cells(f"C{L}:H{L}")
    for col in "ABCDEFGH": ins[f"{col}{L}"].border=box
    ins.row_dimensions[L].height=32; L+=1
L+=1

sec(L,"O DIA A DIA"); L+=1
for perg,resp in [
 ("Concluiu?","Escreva a data em CONCLUÍDA EM. A situação muda sozinha e a "
  "célula fica verde."),
 ("Atrasou?","Escreva o porquê em MOTIVO DO ATRASO / DA MUDANÇA. A lista já traz os "
  "motivos mais comuns, e você pode escrever outro."),
 ("Vai empurrar para outra semana?","Guarde o número da semana antiga em SEMANA ORIG. "
  "e troque a Semana. É contra a original que sai a aderência ao plano original."),
 ("Frota nova?","Digite direto na coluna Frota — a planilha só avisa, não trava. "
  "Depois acrescente o nome na aba Listas para virar opção."),
 ("Entrou fora do plano?","Marque EXTRA na coluna Origem.")]:
    ins[f"A{L}"]=perg; ins[f"A{L}"].font=F(bold=True,size=10,color=NAVY)
    ins[f"A{L}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
    ins.merge_cells(f"A{L}:B{L}")
    ins[f"C{L}"]=resp; ins[f"C{L}"].font=F(size=10)
    ins[f"C{L}"].alignment=Alignment(vertical="top",wrap_text=True,indent=1)
    ins.merge_cells(f"C{L}:H{L}")
    for col in "ABCDEFGH": ins[f"{col}{L}"].border=box
    ins.row_dimensions[L].height=30; L+=1
L+=1

sec(L,"AS CORES — a cor fica na CÉLULA, não na linha inteira"); L+=1
par(L,"Quem muda de cor é a coluna SITUAÇÃO. O resto da linha continua branco, "
      "para o texto da atividade e a observação se lerem. Fora dela só se "
      "pintam três coisas: a coluna ORIGEM quando é EXTRA, a data de conclusão "
      "assim que é escrita, e o motivo quando é preenchido.",alt=32,cor=T2); L+=2
for cor_,nome,txt in [(RU_V,"VENCIDA","o dia dela passou e não foi concluída"),
      (EX_V,"Fecha hoje","é o último dia da janela — tem de sair hoje"),
      (AZ_V,"Em execução","está dentro da janela de execução dela"),
      (OK_V,"Concluída","tem data de conclusão"),
      (AL_V,"Concluída com atraso","saiu depois do dia programado"),
      (AZ_V,"Em programação","você marcou: ainda está sendo encaixada"),
      ("FFEDF2FC","Programada","tem semana e dia, e a janela ainda não abriu"),
      ("FFF2F4F8","Na carteira","ainda sem semana e sem dia"),
      ("FFF6E8CE","Falta a atividade","linha com frota e sem atividade — ela "
       "não conta em lugar nenhum até você escrever o que vai ser feito"),
      ("FFF6E8CE","Falta a semana","tem DIA e não tem SEMANA. Sem semana não "
       "há data, e sem data a atividade some da conta — escreva o número"),
      ("FFE6E9EF","Cancelada","texto riscado; sai dos dois lados da aderência"),
      ("FFFFC97A","EXTRA — coluna Origem","entrou fora do plano; moldura grossa")]:
    c=ins[f"A{L}"]; c.value=nome; c.fill=fill(cor_); c.border=box
    c.font=F(size=10,bold=True); c.alignment=Alignment(horizontal="center")
    ins.merge_cells(f"A{L}:B{L}")
    ins[f"C{L}"]=txt; ins[f"C{L}"].font=F(size=10)
    ins[f"C{L}"].alignment=Alignment(vertical="center",indent=1)
    ins.merge_cells(f"C{L}:H{L}")
    for col in "ABCDEFGH": ins[f"{col}{L}"].border=box
    ins.row_dimensions[L].height=18; L+=1
L+=1
par(L,"As colunas de digitar seguem quatro faixas coloridas, na ordem em que se usam: "
      "o que é o serviço · quem faz · programar · só quando acontecer. "
      "Cinza é calculado — não digite.",alt=30,cor=T2)
ins.column_dimensions["A"].width=14
for col in "BCDEFGH": ins.column_dimensions[col].width=17
ins.sheet_view.showGridLines=False

# ═══════════════════════════════════ IMPRESSÃO E FECHO
for ws,orient,tr in ((pg,"landscape",3),(sm,"landscape",LCAB),
                     (ad,"portrait",None),(ins,"portrait",None),
                     (hj,"landscape",None)):
    ws.page_setup.orientation=orient
    ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.print_options.horizontalCentered=True
    ws.page_margins.left=ws.page_margins.right=0.4
    ws.page_margins.top=ws.page_margins.bottom=0.5
    if tr: ws.print_title_rows=f"1:{tr}"
pg.print_area=f"A1:U{PRIM+max(len(LINHAS),60)+5}"
ad.print_area=f"A1:M{L0+max(N_EX,N_FR)+len(TIPO_LISTA)+12}"   # O:Q são os números dos gráficos
sm.print_area=f"A1:N{LDIA+3}"
hj.print_area=f"A1:I{L2}"
# As onze colunas de cálculo ficam agrupadas e recolhidas: a Programação passa
# de trinta colunas na tela para dezenove, que são as que se digitam. O sinal
# de + na régua abre o grupo quando ele quiser conferir a conta.
pg.column_dimensions.group("V","AI", outline_level=1, hidden=True)
wb.calculation.fullCalcOnLoad=True
for ws in wb.worksheets: ws.sheet_properties.tabColor=NAVY[2:]
# O dia primeiro: é por ele que se começa a manhã.
wb.move_sheet("Hoje", offset=-(wb.sheetnames.index("Hoje")))
wb.active=0
wb.save(SAIDA)
print("planilha montada:",SAIDA)
print("  atividades:",len(LINHAS),"| executantes:",len(EXEC_LISTA),"| frotas:",len(FROTA_LISTA))
from collections import Counter
print("  dias previstos:",dict(Counter(l["dias"] for l in LINHAS)))
